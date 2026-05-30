from flask import Flask, render_template, request, jsonify, session, send_file, Response, send_from_directory
import json
import os
import sys
import threading
import uuid
import re
from datetime import datetime
from log import Logger
from script.public import download_latest_image

try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import load_workbook

import os
print("当前工作目录:", os.getcwd())
print("app.py所在目录:", os.path.dirname(os.path.abspath(__file__)))

# 添加脚本目录到Python路径
# sys.path.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'script'))

log = Logger()

app = Flask(__name__, static_folder='.', static_url_path='/')


app.secret_key = 'your-secret-key'  # 设置secret key用于session加密

@app.errorhandler(Exception)
def handle_exception(e):
    error_msg = str(e)
    # 移除无法编码的特殊字符，避免GBK编码错误
    error_msg = error_msg.encode('utf-8', errors='replace').decode('utf-8')
    log.error(f"全局异常处理: {error_msg}")
    return jsonify({'success': False, 'error': error_msg})

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_MODEL = 'MS'


def resolve_model_file(base_dir, model, filename):
    """优先读取按机型目录文件，不存在则回退到旧路径。"""
    normalized_model = (model or '').strip()
    if normalized_model:
        model_path = os.path.join(BASE_DIR, base_dir, normalized_model, filename)
        if os.path.exists(model_path):
            return model_path
    # 旧目录回退，兼容历史文件布局
    return os.path.join(BASE_DIR, base_dir, filename)


def load_manual_tests(model):
    manual_tests_path = resolve_model_file('config', model, 'manual_tests.json')
    if not os.path.exists(manual_tests_path):
        return []
    with open(manual_tests_path, 'r', encoding='utf-8') as f:
        manual_tests_config = json.load(f)
    manual_tests = manual_tests_config.get('manual_tests', [])
    # 处理 description 字段：如果是数组，则用换行符连接成字符串
    for test in manual_tests:
        if 'description' in test and isinstance(test['description'], list):
            test['description'] = '\n'.join(test['description'])
        # 确保 confirm_content 字段存在
        if 'confirm_content' not in test:
            test['confirm_content'] = ''
    return manual_tests


MS_VERSION_ITEMS = ['pilot_version', 'compass_version', 'rcc_base_version', 'mirror_system']
MR_VERSION_ITEMS = ['pilot_version', 'compass_version', 'mirror_system', 'rws_version']
TW_VERSION_ITEMS = ['pilot_version', 'rcc_base_version', 'robot_version', 'rws_version', 'mirror_system', 'youiscript_version', 'mos_version']
HSR_VERSION_ITEMS = ['pilot_version', 'compass_version', 'mirror_system', 'mos_version_hsr', 'mirror_system_hsr', 'rcc_base_version']

MS_VERSION_STANDARDS = {
    "pilot_version": "release/v3.8.0_xjyw_v1.5.21_20260313",
    "compass_version": "YOUICompassSetup-4.7.4-xjyw-V4.2.0-20260427",
    "rcc_base_version": "2.01",
    "mirror_system": "Pilot-4.0.3-desktop-2025-7-31.img",
}

MR_VERSION_STANDARDS = {
    "pilot_version": "YouiPilot-release_v3.5.2_IDC——10-linux-amd64-20250729.deb",
    "compass_version": "YOUICompass-4.6.0-IDC03-20250225.zip",
    "mirror_system": "Pilot3.7.0_2022-5-18--img",
    "rws_version": "YOUIRWS_v2.0.2"
}

HSR_VERSION_STANDARDS = {
    "pilot_version": "v4.0.6_ex100.251028.1",
    "compass_version": "YOUICompassSetup-4.7.4-xjyw-V3.0-bl-20251229",
    "mirror_system": "Pilot-4.0.3-desktop-2025-7-31.img",
    "mos_version_hsr": "2.6.21_ARIS",
    "mirror_system_hsr": "Pilot-4.0.3-desktop-2025-11-27.img",
    "rcc_base_version": "2.01",
}

TW_VERSION_STANDARDS = {
    "pilot_version": "feature_v4.0.3_xjyw.260413.2-x86_64-running",
    "rcc_base_version": "2.01",
    "robot_version": "robot-v1.1.2",
    "rws_version": "v2.0.2",
    "mirror_system": "Pilot-4.0.3-desktop-2025-7-31.img",
    "youiscript_version": "release_v2.0.0.8-x86_64-running",
    "mos_version": "2.8.5_20260411",
}



# 读取指定IP的测试结果文件
def read_result_file(ip):
    # 确保test_record目录存在
    os.makedirs('test_record', exist_ok=True)
    filename = os.path.join('test_record', f"{ip}.json")
    if not os.path.exists(filename):
        return {}
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"读取{filename}文件失败: {e}")
        return {}

# 写入指定IP的测试结果文件
def write_result_file(ip, data):
    # 确保test_record目录存在
    os.makedirs('test_record', exist_ok=True)
    filename = os.path.join('test_record', f"{ip}.json")
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"写入{filename}文件失败: {e}")
        return False

# 存储测试状态
test_status = {
    'version': 'pending',  # pending, success, failed
    'sensor': 'pending',
    'anti_collision': 'pending',
    'speaker': 'pending',
    'button': 'pending',
    'light': 'pending',
    'ping': 'pending',
    'lift_motor': 'pending',
    'dynamic': 'pending',
    'integrated': 'pending',
    'manual': 'pending'
}

# 存储InspectionAutomation实例，以IP地址为键
dynamic_inspectors = {}

# 存储LidarFieldReader实例，以IP地址为键
lidar_readers = {}

dynamic_workers = {}
dynamic_worker_lock = threading.Lock()

# 存储集成测试实例，以IP地址为键
integrated_workers = {}
integrated_worker_lock = threading.Lock()

# 集成测试子任务描述
integrated_task_descriptions = {
    'light_photo': '可见光拍照',
    'light_video': '可见光录像',
    'thermal_photo': '热成像拍照',
    'thermal_video': '热成像录像',
    'light_PTZ': '云台动作',
    'light_point': '云台移动变倍至预置点10',
    'thermal_point': '云台移动变倍至预置点11',
    'Manual_focusing': '云台变成手动聚焦模式',
    'Auto_focusing': '云台变成自动聚焦模式',
    'thermal_Temperature': '测温任务'
}



# 版本标准配置：定义不同机型的标准版本号
version_standards = {
    'MS': dict(MS_VERSION_STANDARDS),
    'MR': dict(MR_VERSION_STANDARDS),
    'X310': dict(MS_VERSION_STANDARDS),
    'X320': dict(MS_VERSION_STANDARDS),
    'TW': dict(TW_VERSION_STANDARDS),
    'HSR': dict(HSR_VERSION_STANDARDS),
}

# 机型配置字典：定义不同机型对应的测试项和子项
model_config = {
    'MS': {
        'test_items': ['version', 'sensor', 'ping', 'button', 'anti_collision', 'speaker', 'light', 'dynamic', 'manual'],
        'button': ['emergency_stop'],
        'light': ['red', 'blue', 'green'],
        'anti_collision': ['front', 'back'],
        'version': list(MS_VERSION_ITEMS),
        'sensor': ['odom', 'imu_data', 'ks114_sensor', 'encoder', 'scan_1', 'cpu_hz'],
        'ping': ['192.168.0.8', '192.168.2.63', '192.168.2.250', '192.168.0.100', '192.168.2.2', '192.168.0.100:8081']
    },
    'MR': {
        'test_items': ['version', 'sensor', 'ping', 'button', 'anti_collision', 'speaker', 'lift_motor', 'dynamic', 'manual'],
        'button': ['left_emergency_stop', 'right_emergency_stop', 'voice'],
        'light': [],
        'anti_collision': ['front', 'back'],
        'version': list(MR_VERSION_ITEMS),
        'sensor': ['odom', 'imu_data', 'ks114_sensor', 'tfmini_sensor', 'encoder', 'scan_1', 'cpu_hz', 'byz06_sensor', 'fs00802_sensor'],
        'ping': ['192.168.0.8', '192.168.2.63', '192.168.2.250', '192.168.0.100', '192.168.2.2', '192.168.0.100:8081', '192.168.0.50']
    },
    'X310': {
        'test_items': ['version', 'sensor', 'ping', 'button', 'anti_collision', 'speaker', 'light', 'dynamic', 'manual'],
        'button': ['emergency_stop', 'voice'],
        'light': ['red', 'blue', 'green'],
        'anti_collision': ['front', 'back'],
        'version': list(MS_VERSION_ITEMS),
        'sensor': ['odom', 'imu_data', 'ks114_sensor', 'tfmini_sensor', 'encoder', 'scan_1', 'cpu_hz'],
        'ping': ['192.168.0.8', '192.168.2.63', '192.168.2.250', '192.168.0.100', '192.168.2.2', '192.168.0.100:8081']
    },
    'X320': {
        'test_items': ['version', 'sensor', 'ping', 'button', 'anti_collision', 'speaker', 'light', 'dynamic', 'manual'],
        'button': ['emergency_stop', 'voice'],
        'light': ['red', 'blue', 'green'],
        'anti_collision': ['front', 'back'],
        'version': list(MS_VERSION_ITEMS),
        'sensor': ['odom', 'imu_data', 'ks114_sensor', 'tfmini_sensor', 'encoder', 'scan_1', 'cpu_hz'],
        'ping': ['192.168.0.8', '192.168.2.63', '192.168.2.250', '192.168.0.100', '192.168.2.2', '192.168.0.100:8081']
    },
    'TW': {
                'test_items': ['version', 'sensor', 'ping', 'button',  'speaker', 'light', 'integrated', 'manual'],
                'button': ['emergency_stop'],
                'light': ['red', 'blue', 'green', 'front_light', 'back_light', 'charge_relay'],
                'anti_collision': ['front', 'back'],
                'version': list(TW_VERSION_ITEMS),
                'sensor': ['odom', 'imu_data', 'ks114_sensor', 'cpu_hz', 'temperature', 'humidity', 'o2', 'microphone', 'fan_board'],
                'ping': ['192.168.0.8', '192.168.2.63', '192.168.2.250', '192.168.0.100', '192.168.2.2',"192.168.2.100"],
                'integrated': ['light_photo', 'light_video', 'thermal_photo', 'thermal_video', 'thermal_Temperature', 'light_PTZ', 'light_point', 'thermal_point', 'Manual_focusing', 'Auto_focusing']
            },
    'HSR': {
        'test_items': ['version', 'sensor', 'ping', 'button', 'anti_collision', 'speaker', 'light', 'dynamic', 'manual'],
        'button': ['chassis_left_stop', 'chassis_right_stop', 'integrated_left_stop', 'integrated_right_stop', 'unlock_brake'],
        'light': ['red', 'blue', 'green'],
        'anti_collision': ['front', 'back', 'left', 'right'],
        'version': list(HSR_VERSION_ITEMS),
        'sensor': ['odom', 'imu_data', 'encoder', 'scan_1', 'cpu_hz'],
        'ping': ['192.168.0.8', '192.168.2.63', '192.168.2.88', '192.168.2.90', '192.168.2.250', '192.168.2.2', '192.168.2.89', '192.168.0.100:8081']
    }
}


def get_version_fields_by_model(model):
    config = model_config.get(model or '')
    if config and config.get('version'):
        return list(config['version'])
    return list(MS_VERSION_ITEMS)


def build_report_filters(data):
    """构建与前端展示一致的PDF筛选配置。"""
    if not isinstance(data, dict):
        data = {}
    robot_info = data.get('robot_info', {}) if isinstance(data.get('robot_info', {}), dict) else {}
    current_model = robot_info.get('model')
    current_config = model_config.get(current_model, model_config.get('MS', {}))

    filters = {
        'model': current_model,
        'test_items': list(current_config.get('test_items', []))
    }

    for key in ['version', 'sensor', 'ping', 'button', 'light', 'anti_collision', 'integrated']:
        filters[key] = list(current_config.get(key, []))

    return filters

# 测试项顺序列表（默认所有机型都有的测试项）
test_order = ['version', 'sensor', 'ping', 'button', 'anti_collision', 'speaker', 'light', 'lift_motor', 'dynamic', 'integrated', 'manual']



@app.route('/')
def login():
    return render_template('login.html')

@app.route('/main')
def main():
    bound_ip = session.get('bound_ip')
    robot_info = {}
    current_test_status = test_status.copy()
    test_times = {}
    current_model = None
    filtered_test_order = test_order.copy()
    current_button_types = ['emergency_stop', 'voice']
    current_light_types = ['red', 'green', 'blue']
    
    if bound_ip:
        results = read_result_file(bound_ip)
        robot_info = results.get('robot_info', {})
        current_model = robot_info.get('model')
        
        if current_model and current_model in model_config:
            filtered_test_order = model_config[current_model]['test_items']
            current_button_types = model_config[current_model]['button']
            current_light_types = model_config[current_model]['light']
        
        for test_type in filtered_test_order:
            if test_type in results:
                current_test_status[test_type] = results[test_type].get('result', 'pending')
                test_times[test_type] = results[test_type].get('time', '')
            else:
                test_times[test_type] = ''
    
    return render_template('main.html', bound_ip=bound_ip, test_status=current_test_status, active_test=None, robot_info=robot_info, test_times=test_times, test_order=filtered_test_order, model_config=model_config, current_model=current_model, current_button_types=current_button_types, current_light_types=current_light_types)

@app.route('/main/<test_type>')
def test_page(test_type):
    valid_tests = test_order
    bound_ip = session.get('bound_ip')
    robot_info = {}
    current_test_status = test_status.copy()
    test_times = {}
    current_model = None
    filtered_test_order = test_order.copy()
    current_button_types = ['emergency_stop', 'voice']
    current_light_types = ['red', 'green', 'blue']
    
    if bound_ip:
        results = read_result_file(bound_ip)
        robot_info = results.get('robot_info', {})
        current_model = robot_info.get('model')
        
        if current_model and current_model in model_config:
            filtered_test_order = model_config[current_model]['test_items']
            valid_tests = filtered_test_order
            current_button_types = model_config[current_model]['button']
            current_light_types = model_config[current_model]['light']
        
        for t_type in filtered_test_order:
            if t_type in results:
                current_test_status[t_type] = results[t_type].get('result', 'pending')
                test_times[t_type] = results[t_type].get('time', '')
            else:
                test_times[t_type] = ''
    
    if test_type not in valid_tests:
        return render_template('main.html', bound_ip=bound_ip, test_status=current_test_status, active_test=None, robot_info=robot_info, test_times=test_times, test_order=filtered_test_order, model_config=model_config, current_model=current_model, current_button_types=current_button_types, current_light_types=current_light_types)
    return render_template('main.html', bound_ip=bound_ip, test_status=current_test_status, active_test=test_type, robot_info=robot_info, test_times=test_times, test_order=filtered_test_order, model_config=model_config, current_model=current_model, current_button_types=current_button_types, current_light_types=current_light_types)

import subprocess
import platform

def ping_ip(ip):
    """测试IP是否可达，超时1秒"""
    param = '-n' if platform.system().lower() == 'windows' else '-c'
    command = ['ping', param, '1', '-w', '1000', ip]
    try:
        output = subprocess.check_output(command, stderr=subprocess.STDOUT, universal_newlines=True, timeout=1)
        return True
    except subprocess.CalledProcessError:
        return False
    except subprocess.TimeoutExpired:
        return False

@app.route('/bind_ip', methods=['POST'])
def bind_ip():
    ip = request.json.get('ip')
    # 先ping测试IP是否可达
    if not ping_ip(ip):
        return jsonify({'success': False, 'error': 'IP地址不可达，请检查网络连接'})
    
    session['bound_ip'] = ip
    # 读取IP命名的JSON文件并更新robot_info中的ip
    results = read_result_file(ip)
    robot_info = results.get('robot_info', {})
    robot_info['ip'] = ip
    results['robot_info'] = robot_info
    write_result_file(ip, results)
    return jsonify({'success': True, 'bound_ip': ip, 'robot_info': robot_info})

@app.route('/update_robot_info', methods=['POST'])
def update_robot_info():
    sn = request.json.get('sn')
    tester = request.json.get('tester')
    model = request.json.get('model')
    ip = request.json.get('ip')
    
    # 读取IP命名的JSON文件
    results = read_result_file(ip)
    robot_info = {
        'sn': sn,
        'tester': tester,
        'model': model,
        'ip': ip
    }
    results['robot_info'] = robot_info
    write_result_file(ip, results)
    
    return jsonify({'success': True, 'robot_info': robot_info})

@app.route('/unbind_ip', methods=['POST'])
def unbind_ip():
    if 'bound_ip' in session:
        del session['bound_ip']
    return jsonify({'success': True})

@app.route('/get_robot_info', methods=['GET'])
def get_robot_info():
    bound_ip = session.get('bound_ip')
    if not bound_ip:
        return jsonify({'success': False, 'error': '未绑定IP'})
    
    results = read_result_file(bound_ip)
    robot_info = results.get('robot_info', {})
    
    return jsonify({'success': True, 'robot_info': robot_info})

@app.route('/get_test_status', methods=['GET'])
def get_test_status():
    bound_ip = session.get('bound_ip')
    if not bound_ip:
        return jsonify({'success': False, 'error': '未绑定IP'})
    
    results = read_result_file(bound_ip)
    test_results = {}
    test_times = {}
    for test_type in test_order:
        if test_type in results:
            test_results[test_type] = results[test_type].get('result', 'pending')
            test_times[test_type] = results[test_type].get('time', '')
        else:
            test_results[test_type] = 'pending'
            test_times[test_type] = ''
    
    return jsonify({'success': True, 'test_status': test_results, 'test_times': test_times, 'test_order': test_order})

@app.route('/get_model_config', methods=['GET'])
def get_model_config():
    model = request.args.get('model')
    if not model or model not in model_config:
        return jsonify({'success': False, 'error': '无效的机型'})
    return jsonify({'success': True, 'config': model_config[model]})

@app.route('/get_test_order', methods=['GET'])
def get_test_order():
    return jsonify({'success': True, 'test_order': test_order})

@app.route('/update_test_order', methods=['POST'])
def update_test_order():
    new_order = request.json.get('test_order')
    if not new_order or not isinstance(new_order, list):
        return jsonify({'success': False, 'error': '无效的测试顺序数据'})
    
    # 验证所有测试类型是否有效
    valid_test_types = set(['version', 'sensor', 'anti_collision', 'speaker', 'button'])
    if set(new_order) != valid_test_types:
        return jsonify({'success': False, 'error': '测试顺序包含无效的测试类型'})
    
    global test_order
    test_order = new_order
    return jsonify({'success': True, 'test_order': test_order})

@app.route('/get_version_standards', methods=['GET'])
def get_version_standards():
    model = request.args.get('model')
    if not model:
        bound_ip = session.get('bound_ip')
        if bound_ip:
            results = read_result_file(bound_ip)
            robot_info = results.get('robot_info', {})
            model = robot_info.get('model')
    
    if not model:
        return jsonify({'success': False, 'error': '未找到机型信息'})
    
    standards = {}
    
    config_path = os.path.join('config', model, 'version_config.json')
    if os.path.exists(config_path):
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config_data = json.load(f)
                # 复制配置文件中的所有字段，而不是只复制特定字段
                for key, value in config_data.items():
                    standards[key] = value
        except Exception as e:
            log.error(f"读取版本配置文件失败: {str(e)}")
    
    if not standards and model in version_standards:
        standards = version_standards[model]
    
    if not standards:
        return jsonify({'success': False, 'error': '未找到该机型的版本标准'})
    
    return jsonify({'success': True, 'standards': standards, 'model': model})

@app.route('/get_version_info', methods=['GET'])
def get_version_info():
    bound_ip = session.get('bound_ip')
    if not bound_ip:
        return jsonify({'success': False, 'error': '未绑定IP'})
    
    results = read_result_file(bound_ip)
    version_info = results.get('version', {})
    
    return jsonify({'success': True, 'version_info': version_info})

@app.route('/get_hsr_pilot_version', methods=['GET'])
def get_hsr_pilot_version():
    bound_ip = session.get('bound_ip')
    if not bound_ip:
        return jsonify({'success': False, 'error': '未绑定IP'})
    
    try:
        from script.hsr_ssh import HSRSSH
        hsr_ssh = HSRSSH()
        result = hsr_ssh.test_connection(bound_ip)
        return jsonify(result)
    except Exception as e:
        log.error(f"获取HSR Pilot版本失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取版本失败: {str(e)}'})

@app.route('/get_sensor_info', methods=['GET'])
def get_sensor_info():
    bound_ip = session.get('bound_ip')
    if not bound_ip:
        return jsonify({'success': False, 'error': '未绑定IP'})
    
    results = read_result_file(bound_ip)
    sensor_info = results.get('sensor', {})
    
    return jsonify({'success': True, 'sensor_info': sensor_info})



@app.route('/get_anti_collision_info', methods=['GET'])
def get_anti_collision_info():
    bound_ip = session.get('bound_ip')
    if not bound_ip:
        return jsonify({'success': False, 'error': '未绑定IP'})
    
    results = read_result_file(bound_ip)
    anti_collision_info = results.get('anti_collision', {})
    
    # 获取当前机型信息
    robot_info = results.get('robot_info', {})
    current_model = robot_info.get('model', '')
    
    # 根据机型设置默认的防撞条状态
    default_data = {'front': None, 'back': None}
    if current_model == 'HSR':
        default_data['left'] = None
        default_data['right'] = None
    
    return jsonify({'success': True, 'anti_collision_info': anti_collision_info.get('data', default_data), 'newton_info': anti_collision_info.get('newton', {}), 'test_time': anti_collision_info.get('time', '')})

@app.route('/test_anti_collision', methods=['POST'])
def test_anti_collision():
    bound_ip = session.get('bound_ip')
    if not bound_ip:
        return jsonify({'success': False, 'error': '未绑定IP'})
    
    strip_type = request.json.get('strip_type')
    newton_value = request.json.get('newton_value')
    check_only = request.json.get('check_only', False)
    save_result = request.json.get('save_result', False)
    
    if not strip_type:
        return jsonify({'success': False, 'error': '缺少防撞条类型'})
    
    try:
        # 获取当前机型信息
        results = read_result_file(bound_ip)
        robot_info = results.get('robot_info', {})
        current_model = robot_info.get('model', '')
        
        # 使用绑定的IP地址
        log.info(f"防撞条检测 - 机型: {current_model}, IP: {bound_ip}, 类型: {strip_type}")
        
        # 获取防撞条数据
        v = None
        
        # HSR车型的左右防撞条尝试多种获取方式
        if current_model == 'HSR' and (strip_type == 'left' or strip_type == 'right'):
            log.info(f"===== HSR防撞条检测开始 =====")
            log.info(f"机型: {current_model}, IP: {bound_ip}, 类型: {strip_type}")
            
            from script.hsr_bumper import HSRBumperReader
            
            keyword = "leftBumper" if strip_type == 'left' else "rightBumper"
            reader = HSRBumperReader(ip=bound_ip, timeout=2)
            result = reader.get_emergency_stop_status(keyword)
            v = 1 if result else 0
            
            log.info(f"HSR {strip_type}防撞条检测结果: {v}")
            log.info(f"===== HSR防撞条检测结束 =====")
        else:
            # 其他情况使用ROS topic方式
            from script.robot_allstatus import SingleTopicOnceReader
            reader = SingleTopicOnceReader(
                ip=bound_ip,
                topic="/robot/all_status",
                timeout=2
            )
            
            if strip_type == 'front':
                v = reader.get_one_frame("front_safety_edge")
                if v is None:
                    v = reader.get_one_frame("front_edge")
                    if v is None:
                        v = reader.get_one_frame("safety_edge_front")
            elif strip_type == 'back':
                v = reader.get_one_frame("back_safety_edge")
                if v is None:
                    v = reader.get_one_frame("back_edge")
                    if v is None:
                        v = reader.get_one_frame("safety_edge_back")
            elif strip_type == 'left':
                v = reader.get_one_frame("left_safety_edge")
                if v is None:
                    v = reader.get_one_frame("left_edge")
                    if v is None:
                        v = reader.get_one_frame("safety_edge_left")
            elif strip_type == 'right':
                v = reader.get_one_frame("right_safety_edge")
                if v is None:
                    v = reader.get_one_frame("right_edge")
                    if v is None:
                        v = reader.get_one_frame("safety_edge_right")
            else:
                return jsonify({'success': False, 'error': '无效的防撞条类型'})
        
        if v is None:
            log.error(f"防撞条 {strip_type} ({current_model}车型) 未找到对应的数据")
            return jsonify({'success': False, 'error': f'{strip_type}防撞条数据获取失败，请检查配置'})
        
        if check_only:
            # 只返回当前防撞条状态
            return jsonify({'success': True, 'value': v})
        elif save_result:
            # 保存测试结果为成功
            result = "success"
            # 保存测试结果
            results = read_result_file(bound_ip)
            anti_collision_info = results.get('anti_collision', {})
            if 'data' not in anti_collision_info:
                anti_collision_info['data'] = {}
            if 'newton' not in anti_collision_info:
                anti_collision_info['newton'] = {}
            anti_collision_info['data'][strip_type] = result
            if newton_value is not None:
                anti_collision_info['newton'][strip_type] = newton_value
            anti_collision_info['time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            results['anti_collision'] = anti_collision_info
            write_result_file(bound_ip, results)
            return jsonify({'success': True, 'result': result})
        else:
            # 原始测试逻辑（保持兼容）
            if v == 1:
                result = "success"
            else:
                result = "failed"
            # 保存测试结果
            results = read_result_file(bound_ip)
            anti_collision_info = results.get('anti_collision', {})
            if 'data' not in anti_collision_info:
                anti_collision_info['data'] = {}
            if 'newton' not in anti_collision_info:
                anti_collision_info['newton'] = {}
            anti_collision_info['data'][strip_type] = result
            if newton_value is not None:
                anti_collision_info['newton'][strip_type] = newton_value
            anti_collision_info['time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            results['anti_collision'] = anti_collision_info
            write_result_file(bound_ip, results)
            return jsonify({'success': True, 'result': result})
    except Exception as e:
        error_msg = str(e)
        # 移除无法编码的特殊字符，避免GBK编码错误
        error_msg = error_msg.encode('utf-8', errors='replace').decode('utf-8')
        log.error(f"防撞条测试出错: {error_msg}")
        import traceback
        log.error(f"堆栈: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': error_msg})

@app.route('/test_button', methods=['POST'])
def test_button():
    print("========== 测试按钮请求 ==========")
    print(f"请求数据: {request.json}")
    bound_ip = session.get('bound_ip')
    print(f"绑定IP: {bound_ip}")
    if not bound_ip:
        return jsonify({'success': False, 'error': '未绑定IP'})
    
    button_type = request.json.get('button_type')
    check_only = request.json.get('check_only', False)
    save_result = request.json.get('save_result', False)
    print(f"按钮类型: {button_type}, check_only: {check_only}, save_result: {save_result}")
    
    if not button_type:
        return jsonify({'success': False, 'error': '缺少按钮类型'})
    
    try:
        v = None
        if button_type in ['emergency_stop', 'left_emergency_stop', 'right_emergency_stop', 'chassis_left_stop', 'chassis_right_stop', 'integrated_left_stop', 'integrated_right_stop']:
            # 测试急停按钮（包括左急停、右急停、底盘左右急停、上集成左右急停）
            print(f"开始测试急停按钮: {button_type}")
            from script.robot_allstatus import SingleTopicOnceReader
            reader = SingleTopicOnceReader(
                ip=bound_ip,
                topic="/robot/all_status",
                timeout=2
            )
            v = reader.get_one_frame("stop_button")
            print(f"获取到的按钮值: {v}")
        elif button_type == 'voice':
            # 测试语音按钮
            print(f"开始测试语音按钮")
            from script.robot_allstatus import SingleTopicOnceReader
            reader = SingleTopicOnceReader(
                ip=bound_ip,
                topic="status_PLC1",
                timeout=2
            )
            v = reader.get_one_frame("talkback_state")
            print(f"获取到的语音按钮值: {v}")
        elif button_type == 'unlock_brake':
            # 测试解抱闸按钮
            print(f"开始测试解抱闸按钮")
            from script.robot_allstatus import SingleTopicOnceReader
            reader = SingleTopicOnceReader(
                ip=bound_ip,
                topic="/robot/all_status",
                timeout=2
            )
            v = reader.get_one_frame("unlock_brake")
            print(f"获取到的解抱闸按钮值: {v}")
        else:
            return jsonify({'success': False, 'error': '无效的按钮类型'})
        
        if check_only:
            # 只返回当前按钮状态，即使值为None也返回成功状态
            result = {'success': True, 'value': v}
            print(f"check_only模式返回: {result}")
            return jsonify(result)
        elif save_result:
            # 保存测试结果为成功
            result = "success"
            # 保存测试结果
            results = read_result_file(bound_ip)
            button_info = results.get('button', {})
            if 'data' not in button_info:
                button_info['data'] = {}
            button_info['data'][button_type] = result
            button_info['time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            results['button'] = button_info
            write_result_file(bound_ip, results)
            print(f"save_result模式返回成功")
            return jsonify({'success': True, 'result': result})
        else:
            # 原始测试逻辑（保持兼容）
            if v == 1:
                result = "success"
            else:
                result = "failed"
            # 保存测试结果
            results = read_result_file(bound_ip)
            button_info = results.get('button', {})
            if 'data' not in button_info:
                button_info['data'] = {}
            button_info['data'][button_type] = result
            button_info['time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            results['button'] = button_info
            write_result_file(bound_ip, results)
            print(f"普通测试模式返回: {result}")
            return jsonify({'success': True, 'result': result})
    except Exception as e:
        log.error(f"按钮测试出错: {str(e)}")
        import traceback
        error_trace = traceback.format_exc()
        log.error(f"堆栈: {error_trace}")
        print(f"按钮测试异常: {str(e)}")
        print(f"堆栈: {error_trace}")
        # 对于check_only模式，即使出错也返回成功状态，但value为None
        if check_only:
            return jsonify({'success': True, 'value': None})
        return jsonify({'success': False, 'error': f"{str(e)}"})

@app.route('/test_light', methods=['POST'])
def test_light():
    bound_ip = session.get('bound_ip')
    if not bound_ip:
        return jsonify({'success': False, 'error': '未绑定IP'})
    
    light_color = request.json.get('light_color')
    action = request.json.get('action', 'on')  # 默认开启
    if not light_color:
        return jsonify({'success': False, 'error': '缺少灯光颜色'})
    
    try:
        # 获取当前机型信息
        results = read_result_file(bound_ip)
        robot_info = results.get('robot_info', {})
        current_model = robot_info.get('model')
        
        # 根据灯光颜色、动作和机型设置对应的AT命令
        if light_color == 'red':
            at_command = 'AT+setLampColor=1,1'
        elif light_color == 'green':
            if current_model == 'TW' or current_model == 'HSR':
                at_command = 'AT+setLampColor=4,1'
            else:
                at_command = 'AT+setLampColor=2,1'
        elif light_color == 'blue':
            if current_model == 'TW' or current_model == 'HSR':
                at_command = 'AT+setLampColor=10,1'
            else:
                at_command = 'AT+setLampColor=3,1'
        elif light_color == 'front_light':
            at_command = 'AT+PDO=1,1' if action == 'on' else 'AT+PDO=1,0'
        elif light_color == 'back_light':
            at_command = 'AT+PDO=7,1' if action == 'on' else 'AT+PDO=7,0'
        elif light_color == 'charge_relay':
            at_command = 'AT+ChargeContact=1' if action == 'on' else 'AT+ChargeContact=0'
        else:
            return jsonify({'success': False, 'error': '无效的灯光颜色'})
        
        # 调用Ligit类执行灯光控制命令
        from script.at_command import Ligit
        ssh_robot = Ligit()
        ssh_robot.execute_lighting(bound_ip, at_command)
        
        # 总是返回成功，因为测试结果由用户在UI界面选择
        return jsonify({'success': True})
    except Exception as e:
        log.error(f"灯光测试出错: {str(e)}")
        import traceback
        log.error(f"堆栈: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': f"{str(e)}"})

@app.route('/get_light_info')
def get_light_info():
    bound_ip = session.get('bound_ip')
    if not bound_ip:
        return jsonify({'success': False, 'error': '未绑定IP'})
    
    try:
        results = read_result_file(bound_ip)
        light_info = results.get('light', {})
        test_time = light_info.get('time', '')
        return jsonify({'success': True, 'light_info': light_info.get('data', {}), 'test_time': test_time})
    except Exception as e:
        log.error(f"获取灯光测试信息出错: {str(e)}")
        import traceback
        log.error(f"堆栈: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': f"{str(e)}"})

@app.route('/save_light_test_result', methods=['POST'])
def save_light_test_result():
    bound_ip = session.get('bound_ip')
    if not bound_ip:
        return jsonify({'success': False, 'error': '未绑定IP'})
    
    light_color = request.json.get('light_color')
    result = request.json.get('result')
    
    if not light_color or not result:
        return jsonify({'success': False, 'error': '缺少灯光颜色或测试结果'})
    
    try:
        # 保存测试结果
        results = read_result_file(bound_ip)
        light_info = results.get('light', {})
        if 'data' not in light_info:
            light_info['data'] = {}
        light_info['data'][light_color] = result
        light_info['time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        results['light'] = light_info
        write_result_file(bound_ip, results)
        
        return jsonify({'success': True})
    except Exception as e:
        log.error(f"保存灯光测试结果出错: {str(e)}")
        import traceback
        log.error(f"堆栈: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': f"{str(e)}"})

@app.route('/get_button_info', methods=['GET'])
def get_button_info():
    bound_ip = session.get('bound_ip')
    if not bound_ip:
        return jsonify({'success': False, 'error': '未绑定IP'})
    
    results = read_result_file(bound_ip)
    button_info = results.get('button', {})
    
    # 构建返回的按钮测试信息
    info = {
        'emergency_stop': button_info.get('data', {}).get('emergency_stop'),
        'left_emergency_stop': button_info.get('data', {}).get('left_emergency_stop'),
        'right_emergency_stop': button_info.get('data', {}).get('right_emergency_stop'),
        'chassis_left_stop': button_info.get('data', {}).get('chassis_left_stop'),
        'chassis_right_stop': button_info.get('data', {}).get('chassis_right_stop'),
        'integrated_left_stop': button_info.get('data', {}).get('integrated_left_stop'),
        'integrated_right_stop': button_info.get('data', {}).get('integrated_right_stop'),
        'unlock_brake': button_info.get('data', {}).get('unlock_brake'),
        'voice': button_info.get('data', {}).get('voice'),
        'time': button_info.get('time', '')
    }
    
    return jsonify({'success': True, 'button_info': info})

@app.route('/get_ping_info', methods=['GET'])
def get_ping_info():
    bound_ip = session.get('bound_ip')
    if not bound_ip:
        return jsonify({'success': False, 'error': '未绑定IP'})
    
    results = read_result_file(bound_ip)
    ping_info = results.get('ping', {})
    
    return jsonify({'success': True, 'ping_info': ping_info.get('data', {}), 'test_time': ping_info.get('time', '')})

@app.route('/get_dynamic_info', methods=['GET'])
def get_dynamic_info():
    bound_ip = session.get('bound_ip')
    if not bound_ip:
        return jsonify({'success': False, 'error': '未绑定IP'})
    
    # 从script.pdf模块导入动态测试子任务描述
    from script.pdf import dynamic_task_descriptions
    
    # 从结果文件中读取测试内容
    results = read_result_file(bound_ip)
    dynamic_info = results.get('dynamic', {})
    robot_info = results.get('robot_info', {})
    current_model = robot_info.get('model')
    
    # 构建任务状态和描述
    task_status = dynamic_info.get('data', {})
    if not task_status:
        # 根据机型返回不同的默认任务列表
        if current_model == 'HSR':
            task_status = {
                '直线': '未测试', 
                '切区': '未测试', 
                '横移': '未测试', 
                '沟壑': '未测试', 
                '45°夹角': '未测试', 
                '精定位': '未测试', 
                '云台': '未测试'
            }
        else:
            task_status = {'直线': '未测试', '曲线': '未测试', '沟壑': '未测试', '切区': '未测试', '云台': '未测试'}
    
    # 从文件中读取并返回测试内容
    return jsonify({
        'success': True, 
        'task_status': task_status, 
        'test_time': dynamic_info.get('time', ''),
        'task_descriptions': dynamic_task_descriptions,
        'current_step': dynamic_info.get('current_step', ''),
        'step_status': dynamic_info.get('step_status', ''),
        'error': dynamic_info.get('error', ''),
        'result': dynamic_info.get('result', ''),
        'init_steps': dynamic_info.get('init_steps', []),
        'robot_model': current_model
    })


@app.route('/get_manual_info', methods=['GET'])
def get_manual_info():
    bound_ip = session.get('bound_ip')
    if not bound_ip:
        return jsonify({'success': False, 'error': '未绑定IP'})
    
    # 从结果文件中读取测试内容
    results = read_result_file(bound_ip)
    manual_info = results.get('manual', {})
    robot_info = results.get('robot_info', {})
    current_model = robot_info.get('model') or DEFAULT_MODEL
    
    # 从配置文件获取人工测试项
    manual_items_with_desc = load_manual_tests(current_model)
    manual_items = [item['name'] for item in manual_items_with_desc]
    
    # 如果没有人工测试结果，返回默认状态
    if not manual_info:
        return jsonify({
            'success': True,
            'task_status': {item['name']: 'pending' for item in manual_items_with_desc},
            'task_time': {item['name']: '' for item in manual_items_with_desc},
            'task_notes': {item['name']: '' for item in manual_items_with_desc},
            'items': manual_items_with_desc
        })
    
    raw_data = manual_info.get('data', {}) or {}
    task_status = {}
    task_time = {}
    task_notes = {}
    for item in manual_items_with_desc:
        name = item.get('name')
        value = raw_data.get(name)
        if isinstance(value, dict):
            task_status[name] = value.get('result') or value.get('status') or value.get('value') or 'pending'
            task_time[name] = value.get('time') or ''
            task_notes[name] = value.get('note') or ''
        else:
            task_status[name] = value or 'pending'
            task_time[name] = ''
            task_notes[name] = ''
    
    for k, v in raw_data.items():
        if k in task_status:
            continue
        if isinstance(v, dict):
            task_status[k] = v.get('result') or v.get('status') or v.get('value') or 'pending'
            task_time[k] = v.get('time') or ''
            task_notes[k] = v.get('note') or ''
        else:
            task_status[k] = v or 'pending'
            task_time[k] = ''
            task_notes[k] = ''
    
    return jsonify({
        'success': True,
        'task_status': task_status,
        'task_time': task_time,
        'task_notes': task_notes,
        'items': manual_items_with_desc,
        'test_time': manual_info.get('time', '')
    })


@app.route('/get_integrated_info', methods=['GET'])
def get_integrated_info():
    bound_ip = session.get('bound_ip')
    if not bound_ip:
        return jsonify({'success': False, 'error': '未绑定IP'})
    
    # 从结果文件中读取测试内容
    results = read_result_file(bound_ip)
    integrated_info = results.get('integrated', {})
    robot_info = results.get('robot_info', {})
    current_model = robot_info.get('model') or DEFAULT_MODEL
    
    # 获取集成测试子任务
    if current_model in model_config and 'integrated' in model_config[current_model]:
        integrated_items = model_config[current_model]['integrated']
    else:
        integrated_items = []
    
    # 构建任务状态和描述
    task_status = integrated_info.get('data', {})
    if not task_status:
        task_status = {item: 'pending' for item in integrated_items}
    
    # 从文件中读取并返回测试内容
    return jsonify({
        'success': True, 
        'task_status': task_status, 
        'test_time': integrated_info.get('time', ''),
        'task_descriptions': integrated_task_descriptions,
        'error': integrated_info.get('error', ''),
        'result': integrated_info.get('result', ''),
        'current_task': integrated_info.get('current_task', '')
    })


@app.route('/test_single_integrated_task', methods=['POST'])
def test_single_integrated_task():
    bound_ip = session.get('bound_ip')
    if not bound_ip:
        return jsonify({'success': False, 'error': '未绑定IP'})
    
    task_name = request.json.get('task_name')
    if not task_name:
        return jsonify({'success': False, 'error': '缺少任务名称'})
    
    try:
        from script.tw_mos import MosPTZController
        from script.public import download_mos_file
        
        # 任务与文件名的映射
        task_filename_map = {
            'light_photo': '可见光拍照.jpg',
            'light_video': '可见光录像.mp4',
            'thermal_photo': '热成像拍照.jpg',
            'thermal_video': '热成像录像.mp4'
        }
        
        # 执行MOS任务
        controller = MosPTZController(bound_ip)
        response = controller.execute_and_poll_status(task_name)
        log.info(f"单独任务 {task_name} 执行结果: {response}")
        
        # 判断任务是否成功
        task_success = False
        
        # 对于需要检查路径的任务，必须获取到路径或数据才认为成功
        path_check_tasks = ['light_video', 'thermal_photo', 'thermal_video', 'light_photo', 'thermal_Temperature']
        if task_name in path_check_tasks:
            if isinstance(response, str) and response:
                task_success = True
            elif isinstance(response, tuple) and len(response) == 2:
                # 温度数据返回的是元组
                task_success = True
        else:
            # 其他任务只要有响应就认为成功
            if response is not None:
                task_success = True
        
        # 如果是需要下载文件的任务，下载文件
        if task_name in task_filename_map and isinstance(response, str):
            local_file = download_mos_file(bound_ip, response, task_filename_map[task_name])
            if local_file:
                log.info(f"文件下载成功: {local_file}")
        
        # 保存结果
        results = read_result_file(bound_ip)
        integrated_info = results.get('integrated', {})
        if 'data' not in integrated_info:
            integrated_info['data'] = {}
        
        if task_success:
            integrated_info['data'][task_name] = 'success'
            results['integrated'] = integrated_info
            write_result_file(bound_ip, results)
            return jsonify({'success': True, 'result': 'success'})
        else:
            integrated_info['data'][task_name] = 'failed'
            results['integrated'] = integrated_info
            write_result_file(bound_ip, results)
            return jsonify({'success': False, 'error': '任务执行失败或超时'})
    except Exception as e:
        log.error(f"单独任务执行失败: {str(e)}")
        import traceback
        log.error(f"堆栈: {traceback.format_exc()}")
        
        # 保存失败结果
        results = read_result_file(bound_ip)
        integrated_info = results.get('integrated', {})
        if 'data' not in integrated_info:
            integrated_info['data'] = {}
        integrated_info['data'][task_name] = 'failed'
        results['integrated'] = integrated_info
        write_result_file(bound_ip, results)
        
        return jsonify({'success': False, 'error': str(e)})


@app.route('/get_lidar_field', methods=['GET'])
def get_lidar_field():
    bound_ip = session.get('bound_ip')
    if not bound_ip:
        return jsonify({'success': False, 'error': '未绑定IP'})
    
    if bound_ip not in lidar_readers:
        try:
            from script.get_laser_filed import LidarFieldReader
            reader = LidarFieldReader(bound_ip, interval=0.2, recv_timeout=0.8)
            lidar_readers[bound_ip] = reader
            reader.start()
        except Exception:
            return jsonify({'success': True, 'data': None})
    
    reader = lidar_readers[bound_ip]
    return jsonify({'success': True, 'data': reader.last_data})



@app.route('/start_test', methods=['POST'])
def start_test():
    test_type = request.json.get('test_type')
    bound_ip = session.get('bound_ip')
    
    # 版本检测的特殊处理
    if test_type == 'version':
        print("开始版本检测...")
        if not bound_ip:
            return jsonify({'success': False, 'error': '请先绑定IP地址'})
        
        try:
            from script.version import MirrorSystemReader

            # 获取当前机型信息
            results = read_result_file(bound_ip)
            robot_info = results.get('robot_info', {})
            current_model = robot_info.get('model')
            if not current_model:
                return jsonify({'success': False, 'error': '未获取到机型信息'})
            
            reader = MirrorSystemReader(model=current_model)
            log.info("MirrorSystemReader创建成功")

            needed_fields = get_version_fields_by_model(current_model)
            result = reader.read_fields(bound_ip, needed_fields)
            log.info(f"读取结果: {result}")
            
            # 自动比较版本值
            test_result = 'success'
            if current_model in version_standards:
                standards = version_standards[current_model]
                for field in needed_fields:
                    if field in standards and field in result:
                        standard_value = standards[field]
                        if standard_value in (None, ''):
                            continue
                        actual_value = result[field]
                        # 检查是否为错误信息
                        if isinstance(actual_value, str) and (actual_value.startswith('❌') or actual_value.startswith('🚨') or actual_value.startswith('⚠️')):
                            test_result = 'failed'
                            break
                        # 比较版本值
                        if str(standard_value) != str(actual_value):
                            test_result = 'failed'
                            break
            
            # 添加测试结果和时间
            result['result'] = test_result
            result['time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            results = read_result_file(bound_ip)
            results['version'] = result
            write_result_file(bound_ip, results)
            
            return jsonify({'success': True, 'version_info': result})
        except Exception as e:
            log.error(f"错误: {str(e)}")
            import traceback
            log.error(f"堆栈: {traceback.format_exc()}")
            return jsonify({'success': False, 'error': f"{str(e)}"})
    
    # 传感器检测的特殊处理
    elif test_type == 'sensor':
        log.info("开始传感器检测...")
        if not bound_ip:
            return jsonify({'success': False, 'error': '请先绑定IP地址'})
        
        try:
            from script.sensor_topic import RosTopicReader
            # 获取当前机型信息
            results = read_result_file(bound_ip)
            robot_info = results.get('robot_info', {})
            current_model = robot_info.get('model')
            reader = RosTopicReader(bound_ip, model=current_model)
            log.info("RosTopicReader创建成功")
            sensor_results = reader.get_all_data()
            log.info(f"传感器检测结果: {sensor_results}")
            
            # 检查是否为NaN的辅助函数
            def is_nan(value):
                """检查值是否为NaN"""
                import math
                try:
                    return math.isnan(float(value))
                except:
                    return False
            
            # 获取当前机型配置的传感器列表
            model_sensors = []
            if current_model in model_config and 'sensor' in model_config[current_model]:
                model_sensors = model_config[current_model]['sensor']
            
            # 传感器状态判断 - 只检查配置中指定的传感器
            test_result = 'success'
            for topic, data in sensor_results.items():
                # 只检查配置中指定的传感器
                if model_sensors and topic not in model_sensors:
                    continue
                    
                if data is None or data == '无数据':
                    test_result = 'failed'
                    log.warning(f"传感器 {topic} 无数据")
                    break
                
                # 检查数据中是否包含NaN
                has_nan = False
                if isinstance(data, list):
                    for item in data:
                        if is_nan(item):
                            has_nan = True
                            break
                elif isinstance(data, dict):
                    for key, value in data.items():
                        if is_nan(value):
                            has_nan = True
                            break
                else:
                    if is_nan(data):
                        has_nan = True
                
                if has_nan:
                    test_result = 'failed'
                    log.warning(f"传感器 {topic} 包含NaN")
                    break
                
                # 对cpu_hz的特殊处理：HSR机型阈值为2400
                if topic == 'cpu_hz' and isinstance(data, list):
                    min_hz = 2400 if current_model == 'HSR' else 2400
                    for hz in data:
                        if hz < min_hz:
                            log.warning(f"传感器 cpu_hz 值 {hz} 低于阈值 {min_hz}")
                            test_result = 'failed'
                            break
                    if test_result == 'failed':
                        break
                # 对ks114_sensor的特殊处理：每个元素大于0才算成功（仅非HSR机型）
                if topic == 'ks114_sensor' and isinstance(data, list) and current_model != 'HSR':
                    for value in data:
                        if value <= 5:
                            log.warning(f"传感器 ks114_sensor 值 {value} <= 5")
                            test_result = 'failed'
                            break
                    if test_result == 'failed':
                        break
                # 对microphone的特殊处理：必须为True才算成功（仅TW机型）
                if topic == 'microphone' and current_model == 'TW':
                    if data is not True:
                        log.warning(f"传感器 microphone 不是True")
                        test_result = 'failed'
                        break
            
            # 保存传感器检测结果
            results = read_result_file(bound_ip)
            results['sensor'] = {
                'result': test_result,
                'data': sensor_results,
                'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            write_result_file(bound_ip, results)
            
            return jsonify({'success': True, 'sensor_info': sensor_results, 'result': test_result})
        except Exception as e:
            log.error(f"错误: {str(e)}")      
            import traceback
            log.error(f"堆栈: {traceback.format_exc()}")
            return jsonify({'success': False, 'error': f"{str(e)}"})
    
    # 扬声器检测的特殊处理
    elif test_type == 'speaker':
        log.info("开始扬声器检测...")
        if not bound_ip:
            return jsonify({'success': False, 'error': '请先绑定IP地址'})
        
        try:
            from script.play_speaker import RemoteDesktopVolumeBell
            player = RemoteDesktopVolumeBell(ip=bound_ip)
            log.info("RemoteDesktopVolumeBell创建成功")
            player.play()
            
            # 保存扬声器检测结果
            results = read_result_file(bound_ip)
            results['speaker'] = {
                'result': 'pending',  # 初始状态为待完成
                'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            write_result_file(bound_ip, results)
            
            return jsonify({'success': True})
        except Exception as e:
            log.error(f"错误: {str(e)}")      
            import traceback
            log.error(f"堆栈: {traceback.format_exc()}")
            return jsonify({'success': False, 'error': f"{str(e)}"})
    
    # 网络Ping测试的特殊处理
    elif test_type == 'ping':
        log.info("开始网络Ping测试...")
        if not bound_ip:
            return jsonify({'success': False, 'error': '请先绑定IP地址'})
        
        try:
            from script.ip import PingManager
            # 执行ping测试，使用示例中的IP配置
            # 获取当前机型信息
            results = read_result_file(bound_ip)
            robot_info = results.get('robot_info', {})
            current_model = robot_info.get('model')
            
            # 根据机型选择不同的ping目标
            if current_model in model_config and 'ping' in model_config[current_model]:
                ping_targets = model_config[current_model]['ping']
            else:
                ping_targets = ["192.168.0.8", "192.168.2.63", "192.168.2.250", "192.168.0.100", "192.168.2.2", "192.168.0.100:8081"]
            
            # HSR机型：Compass接口(192.168.0.100:8081)不需要实际测试，直接返回固定的0.1ms
            compass_target = "192.168.0.100:8081"
            skip_targets = []
            if current_model == 'HSR' and compass_target in ping_targets:
                skip_targets.append(compass_target)
                ping_targets = [t for t in ping_targets if t != compass_target]
            
            manager = PingManager(
                ssh_ip=bound_ip,
                targets=ping_targets,
            )
            log.info("PingManager创建成功")
            ping_results = manager.run_all()
            log.info(f"Ping测试结果: {ping_results}")
            
            # 对于跳过的目标，直接设置固定延迟值
            for target in skip_targets:
                ping_results[target] = "0.100 ms"
                log.info(f"跳过测试 {target}，返回固定延迟: 0.100 ms")
            
            # Ping标准值（只要小于这些值就行，因为延迟越小越好）
            ping_standard = {}
            for ip in ping_targets:
                ping_standard[ip] = "100 ms"
            
            # 为跳过的目标也添加标准值
            for target in skip_targets:
                ping_standard[target] = "100 ms"
            
            # 自动比较Ping值
            test_result = 'success'
            for ip, standard_value in ping_standard.items():
                if ip in ping_results:
                    actual_value = ping_results[ip]
                    if actual_value:
                        # 提取数值部分
                        try:
                            standard_num = float(standard_value.replace(' ms', ''))
                            actual_num = float(actual_value.replace(' ms', ''))
                            if actual_num > standard_num:
                                test_result = 'failed'
                                break
                        except ValueError:
                            test_result = 'failed'
                            break
                    # 如果结果是null，算成功
            
            # 保存Ping测试结果
            results = read_result_file(bound_ip)
            results['ping'] = {
                'result': test_result,
                'data': ping_results,
                'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            write_result_file(bound_ip, results)
            
            return jsonify({'success': True, 'ping_info': ping_results, 'result': test_result})
        except Exception as e:
            log.error(f"错误: {str(e)}")      
            import traceback
            log.error(f"堆栈: {traceback.format_exc()}")
            return jsonify({'success': False, 'error': f"{str(e)}"})
    
    # 升降电机测试的特殊处理
    elif test_type == 'lift_motor':
        log.info("开始升降电机测试...")
        if not bound_ip:
            return jsonify({'success': False, 'error': '请先绑定IP地址'})
        
        try:
            from script.lift_motor import LiftingPlatformController
            # 执行升降电机测试
            controller = LiftingPlatformController(bound_ip)
            controller.run()
            log.info("升降电机测试执行成功")
            
            # 保存升降电机测试结果
            results = read_result_file(bound_ip)
            results['lift_motor'] = {
                'result': 'pending',  # 初始状态为待完成
                'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            write_result_file(bound_ip, results)
            
            return jsonify({'success': True})
        except Exception as e:
            log.error(f"错误: {str(e)}")      
            import traceback
            log.error(f"堆栈: {traceback.format_exc()}")
            return jsonify({'success': False, 'error': f"{str(e)}"})
    
    # 动态测试的特殊处理
    elif test_type == 'dynamic':
        log.info("开始动态测试...")
        if not bound_ip:
            return jsonify({'success': False, 'error': '请先绑定IP地址'})
        
        try:
            from script.dynamic import InspectionAutomation
            from script.get_laser_filed import LidarFieldReader
            results = read_result_file(bound_ip)
            robot_info = results.get('robot_info', {})
            current_model = robot_info.get('model') or DEFAULT_MODEL

            old_worker = None
            with dynamic_worker_lock:
                old_worker = dynamic_workers.get(bound_ip)
                if old_worker:
                    old_worker['stop_event'].set()

            if old_worker:
                old_thread = old_worker.get('thread')
                if old_thread and old_thread.is_alive():
                    old_thread.join(timeout=2)

            if bound_ip not in lidar_readers:
                reader = LidarFieldReader(bound_ip, interval=0.2, recv_timeout=0.8)
                lidar_readers[bound_ip] = reader
                reader.start()

            inspector = InspectionAutomation(bound_ip, current_model)
            dynamic_inspectors[bound_ip] = inspector
            
            # 根据机型设置不同的初始状态
            if current_model == 'HSR':
                initial_status = {
                    '直线': 'failed',
                    '切区': 'failed',
                    '横移': 'failed',
                    '沟壑': 'failed',
                    '45°夹角': 'failed',
                    '精定位': 'failed',
                    '云台': 'failed'
                }
            else:
                initial_status = {
                    '直线': 'failed',
                    '曲线': 'failed',
                    '沟壑': 'failed',
                    '切区': 'failed',
                    '云台': 'failed'
                }
            inspector.save_task_status(initial_status, current_step="准备执行动态任务", step_status="running", error="", result="running")

            run_id = str(uuid.uuid4())
            stop_event = threading.Event()

            def run_tasks_async(run_id_local, stop_event_local):
                try:
                    task_results = inspector.run_tasks(stop_event=stop_event_local)
                    log.info(f"动态测试执行成功: {task_results}")

                    with dynamic_worker_lock:
                        active_worker = dynamic_workers.get(bound_ip)
                        if not active_worker or active_worker.get('run_id') != run_id_local:
                            return

                    results = read_result_file(bound_ip)
                    dynamic_info = results.get('dynamic', {})
                    if not isinstance(dynamic_info, dict):
                        dynamic_info = {}

                    task_data = dynamic_info.get('data', {})
                    if not isinstance(task_data, dict):
                        task_data = {}

                    task_statuses = {k: v for k, v in task_data.items() if k != 'time'}
                    has_error = bool(dynamic_info.get('error')) or bool(task_results.get('error'))
                    final_result = 'success' if task_statuses and all(v == 'success' for v in task_statuses.values()) else 'failed'
                    if has_error:
                        final_result = 'failed'

                    dynamic_info['result'] = final_result
                    dynamic_info['time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    if has_error and not dynamic_info.get('error') and task_results.get('error'):
                        dynamic_info['error'] = task_results.get('error')
                    results['dynamic'] = dynamic_info
                    write_result_file(bound_ip, results)
                except Exception as e:
                    inspector.save_task_status(None, current_step="动态测试异常", step_status="failed", error=str(e), result="failed")
                    log.error(f"错误: {str(e)}")      
                    import traceback
                    log.error(f"堆栈: {traceback.format_exc()}")
                finally:
                    with dynamic_worker_lock:
                        active_worker = dynamic_workers.get(bound_ip)
                        if active_worker and active_worker.get('run_id') == run_id_local:
                            dynamic_workers.pop(bound_ip, None)

            worker_thread = threading.Thread(target=run_tasks_async, args=(run_id, stop_event), daemon=True)
            with dynamic_worker_lock:
                dynamic_workers[bound_ip] = {'run_id': run_id, 'stop_event': stop_event, 'thread': worker_thread}
            worker_thread.start()
            
            return jsonify({'success': True, 'task_status': initial_status, 'current_step': '准备执行动态任务', 'step_status': 'running'})
        except Exception as e:
            log.error(f"错误: {str(e)}")      
            import traceback
            log.error(f"堆栈: {traceback.format_exc()}")
            return jsonify({'success': False, 'error': f"{str(e)}"})
    
    # 集成测试的特殊处理
    elif test_type == 'integrated':
        log.info("开始集成测试...")
        if not bound_ip:
            return jsonify({'success': False, 'error': '请先绑定IP地址'})
        
        try:
            from script.tw_mos import MosPTZController
            results = read_result_file(bound_ip)
            robot_info = results.get('robot_info', {})
            current_model = robot_info.get('model') or DEFAULT_MODEL

            # 停止旧的集成测试
            old_worker = None
            with integrated_worker_lock:
                old_worker = integrated_workers.get(bound_ip)
                if old_worker:
                    old_worker['stop_event'].set()

            if old_worker:
                old_thread = old_worker.get('thread')
                if old_thread and old_thread.is_alive():
                    old_thread.join(timeout=2)

            # 获取集成测试子任务
            if current_model in model_config and 'integrated' in model_config[current_model]:
                integrated_items = model_config[current_model]['integrated']
            else:
                integrated_items = []

            # 初始化状态
            initial_status = {item: 'pending' for item in integrated_items}
            
            # 保存初始状态，第一个任务标记为running
            if integrated_items:
                initial_status[integrated_items[0]] = 'running'
            
            # 保存初始状态
            results = read_result_file(bound_ip)
            integrated_info = {
                'data': initial_status,
                'error': '',
                'result': 'running',
                'time': '',
                'current_task': integrated_items[0] if integrated_items else ''
            }
            results['integrated'] = integrated_info
            write_result_file(bound_ip, results)

            run_id = str(uuid.uuid4())
            stop_event = threading.Event()

            def run_integrated_tasks_async(run_id_local, stop_event_local):
                try:
                    from script.public import download_mos_file
                    # 初始化MOS控制器 - 使用机器人IP
                    controller = MosPTZController(bound_ip)
                    
                    task_statuses = initial_status.copy()
                    all_success = True
                    
                    # 任务与文件名的映射
                    task_filename_map = {
                        'light_photo': '可见光拍照.jpg',
                        'light_video': '可见光录像.mp4',
                        'thermal_photo': '热成像拍照.jpg',
                        'thermal_video': '热成像录像.mp4'
                    }
                    
                    for i, task_name in enumerate(integrated_items):
                        if stop_event_local.is_set():
                            break
                        
                        # 标记当前任务为running
                        task_statuses[task_name] = 'running'
                        results = read_result_file(bound_ip)
                        integrated_info = results.get('integrated', {})
                        integrated_info['data'] = task_statuses.copy()
                        integrated_info['current_task'] = task_name
                        results['integrated'] = integrated_info
                        write_result_file(bound_ip, results)
                        
                        try:
                            # 使用 execute_and_poll_status 执行任务并等待完成
                            response = controller.execute_and_poll_status(task_name)
                            log.info(f"任务 {task_name} 执行结果: {response}")
                            
                            # 如果是需要下载文件的任务，下载文件
                            if task_name in task_filename_map and isinstance(response, str):
                                local_file = download_mos_file(bound_ip, response, task_filename_map[task_name])
                                if local_file:
                                    log.info(f"文件下载成功: {local_file}")
                            
                            # 简单判断：如果有响应则认为成功
                            task_statuses[task_name] = 'success'
                        except Exception as e:
                            log.error(f"任务 {task_name} 执行失败: {str(e)}")
                            task_statuses[task_name] = 'failed'
                            all_success = False
                        
                        # 保存当前任务状态
                        results = read_result_file(bound_ip)
                        integrated_info = results.get('integrated', {})
                        integrated_info['data'] = task_statuses
                        
                        # 更新下一个任务为running（如果有的话）
                        if i + 1 < len(integrated_items):
                            next_task = integrated_items[i + 1]
                            integrated_info['current_task'] = next_task
                        
                        results['integrated'] = integrated_info
                        write_result_file(bound_ip, results)
                        
                        # 任务间短暂延迟
                        import time
                        time.sleep(1)
                    
                    # 所有任务执行完成
                    final_result = 'success' if all_success else 'failed'
                    
                    results = read_result_file(bound_ip)
                    integrated_info = results.get('integrated', {})
                    integrated_info['result'] = final_result
                    integrated_info['time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    results['integrated'] = integrated_info
                    write_result_file(bound_ip, results)
                    
                    log.info(f"集成测试执行完成，结果: {final_result}")
                except Exception as e:
                    results = read_result_file(bound_ip)
                    integrated_info = results.get('integrated', {})
                    integrated_info['error'] = str(e)
                    integrated_info['result'] = 'failed'
                    results['integrated'] = integrated_info
                    write_result_file(bound_ip, results)
                    log.error(f"错误: {str(e)}")      
                    import traceback
                    log.error(f"堆栈: {traceback.format_exc()}")
                finally:
                    with integrated_worker_lock:
                        active_worker = integrated_workers.get(bound_ip)
                        if active_worker and active_worker.get('run_id') == run_id_local:
                            integrated_workers.pop(bound_ip, None)

            worker_thread = threading.Thread(target=run_integrated_tasks_async, args=(run_id, stop_event), daemon=True)
            with integrated_worker_lock:
                integrated_workers[bound_ip] = {'run_id': run_id, 'stop_event': stop_event, 'thread': worker_thread}
            worker_thread.start()
            
            return jsonify({'success': True, 'task_status': initial_status})
        except Exception as e:
            log.error(f"错误: {str(e)}")      
            import traceback
            log.error(f"堆栈: {traceback.format_exc()}")
            return jsonify({'success': False, 'error': f"{str(e)}"})
    
    # 其他测试类型的处理
    return jsonify({'success': True})

@app.route('/end_test', methods=['POST'])
def end_test():
    test_type = request.json.get('test_type')
    result = request.json.get('result')  # success or failed
    global test_status
    if result is not None:
        test_status[test_type] = result
    
    # 保存测试结果到以IP地址命名的json文件
    bound_ip = session.get('bound_ip')
    
    # 如果是动态测试结束，停止激光数据采集
    if test_type == 'dynamic' and bound_ip:
        worker = None
        with dynamic_worker_lock:
            worker = dynamic_workers.get(bound_ip)
            if worker:
                worker['stop_event'].set()
        if worker:
            worker_thread = worker.get('thread')
            if worker_thread and worker_thread.is_alive():
                worker_thread.join(timeout=2)
            with dynamic_worker_lock:
                active_worker = dynamic_workers.get(bound_ip)
                if active_worker and active_worker.get('run_id') == worker.get('run_id'):
                    dynamic_workers.pop(bound_ip, None)
        if bound_ip in lidar_readers:
            reader = lidar_readers[bound_ip]
            reader.stop()
            del lidar_readers[bound_ip]
    
    # 如果是集成测试结束，停止集成测试
    if test_type == 'integrated' and bound_ip:
        worker = None
        with integrated_worker_lock:
            worker = integrated_workers.get(bound_ip)
            if worker:
                worker['stop_event'].set()
        if worker:
            worker_thread = worker.get('thread')
            if worker_thread and worker_thread.is_alive():
                worker_thread.join(timeout=2)
            with integrated_worker_lock:
                active_worker = integrated_workers.get(bound_ip)
                if active_worker and active_worker.get('run_id') == worker.get('run_id'):
                    integrated_workers.pop(bound_ip, None)
    
    if bound_ip:
        # 读取现有结果
        results = read_result_file(bound_ip)
        
        # 获取当前时间
        test_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # 对于版本测试，保留版本信息并添加测试结果和时间
        if test_type == 'version':
            # 保留现有的版本信息
            version_info = results.get('version', {})
            # 更新测试结果和时间
            version_info['result'] = result
            version_info['time'] = test_time
            results['version'] = version_info
        # 对于传感器测试，保留传感器数据并更新测试结果和时间
        elif test_type == 'sensor':
            # 保留现有的传感器数据
            sensor_info = results.get('sensor', {})
            # 更新测试结果和时间
            sensor_info['result'] = result
            sensor_info['time'] = test_time
            results['sensor'] = sensor_info
        # 对于防撞条测试，保留防撞条状态数据并更新测试结果和时间
        elif test_type == 'anti_collision':
            # 保留现有的防撞条状态数据
            anti_collision_info = results.get('anti_collision', {})
            # 确保data字段存在，避免数据丢失
            if 'data' not in anti_collision_info:
                anti_collision_info['data'] = {
                    'front': None,
                    'back': None,
                    'left': None,
                    'right': None
                }
            if 'newton' not in anti_collision_info:
                anti_collision_info['newton'] = {}
            # 保存牛顿值
            newton_values = request.json.get('newton_values', {})
            for strip, newton_val in newton_values.items():
                anti_collision_info['newton'][strip] = newton_val
            # 更新测试结果和时间
            anti_collision_info['result'] = result
            anti_collision_info['time'] = test_time
            results['anti_collision'] = anti_collision_info
        # 对于按钮测试，保留按钮测试数据并更新测试结果和时间
        elif test_type == 'button':
            # 保留现有的按钮测试数据
            button_info = results.get('button', {})
            # 确保data字段存在，避免数据丢失
            if 'data' not in button_info:
                button_info['data'] = {
                    'emergency_stop': None,
                    'left_emergency_stop': None,
                    'right_emergency_stop': None,
                    'chassis_left_stop': None,
                    'chassis_right_stop': None,
                    'integrated_left_stop': None,
                    'integrated_right_stop': None,
                    'unlock_brake': None,
                    'voice': None
                }
            # 更新测试结果和时间
            button_info['result'] = result
            button_info['time'] = test_time
            results['button'] = button_info
        # 对于灯光测试，保留灯光测试数据并更新测试结果和时间
        elif test_type == 'light':
            # 保留现有的灯光测试数据
            light_info = results.get('light', {})
            # 确保data字段存在，避免数据丢失
            if 'data' not in light_info:
                light_info['data'] = {
                    'red': None,
                    'blue': None,
                    'green': None,
                    'front_light': None,
                    'back_light': None
                }
            # 更新测试结果和时间
            light_info['result'] = result
            light_info['time'] = test_time
            results['light'] = light_info
        # 对于ping测试，保留ping测试数据并更新测试结果和时间
        elif test_type == 'ping':
            # 保留现有的ping测试数据
            ping_info = results.get('ping', {})
            # 确保data字段存在，避免数据丢失
            if 'data' not in ping_info:
                ping_info['data'] = {}
            # 更新测试结果和时间
            ping_info['result'] = result
            ping_info['time'] = test_time
            results['ping'] = ping_info
        # 对于动态测试，保留动态测试数据并更新测试结果和时间
        elif test_type == 'dynamic':
            # 保留现有的动态测试数据
            dynamic_info = results.get('dynamic', {})
            # 确保data字段存在，避免数据丢失
            if 'data' not in dynamic_info:
                dynamic_info['data'] = {}
            # 更新测试结果和时间
            dynamic_info['result'] = result
            dynamic_info['time'] = test_time
            results['dynamic'] = dynamic_info
        # 对于集成测试，保留集成测试数据并更新测试结果和时间
        elif test_type == 'integrated':
            # 保留现有的集成测试数据
            integrated_info = results.get('integrated', {})
            # 确保data字段存在，避免数据丢失
            if 'data' not in integrated_info:
                integrated_info['data'] = {}
            # 更新测试结果和时间
            integrated_info['result'] = result
            integrated_info['time'] = test_time
            results['integrated'] = integrated_info
        # 对于人工测试，保存测试数据并更新测试结果和时间
        elif test_type == 'manual':
            manual_payload = request.json.get('data')
            manual_info = results.get('manual', {}) or {}
            existing_data = manual_info.get('data', {}) or {}
            if not isinstance(existing_data, dict):
                existing_data = {}
            
            updated_items = {}
            if isinstance(manual_payload, dict):
                for name, value in manual_payload.items():
                    if isinstance(value, dict):
                        status_value = value.get('result') or value.get('status') or value.get('value')
                        note_value = value.get('note')
                    else:
                        status_value = value
                        note_value = None
                    if status_value is None and note_value is None:
                        continue
                    if name in existing_data and isinstance(existing_data[name], dict):
                        # 保留现有的结果和时间
                        existing_result = existing_data[name].get('result')
                        existing_time = existing_data[name].get('time')
                        existing_note = existing_data[name].get('note')
                        existing_data[name] = {
                            'result': existing_result if status_value is None else status_value,
                            'time': existing_time if status_value is None else test_time
                        }
                        if note_value is not None:
                            existing_data[name]['note'] = note_value
                        elif existing_note is not None:
                            # 保留现有的备注
                            existing_data[name]['note'] = existing_note
                    else:
                        existing_data[name] = {
                            'result': status_value or 'pending',
                            'time': test_time
                        }
                        if note_value is not None:
                            existing_data[name]['note'] = note_value
                    updated_items[name] = existing_data[name]
            
            manual_info['data'] = existing_data
            manual_info['time'] = test_time
            if result is not None:
                manual_info['result'] = result
            results['manual'] = manual_info
        else:
            # 其他测试类型的处理
            results[test_type] = {
                'result': result,
                'time': test_time
            }
        
        # 写回文件
        write_result_file(bound_ip, results)
    
    resp = {'success': True, 'test_status': test_status}
    if bound_ip and test_type == 'manual':
        manual_info = read_result_file(bound_ip).get('manual', {}) or {}
        resp['manual_time'] = manual_info.get('time', '')
        resp['updated_items'] = updated_items if 'updated_items' in locals() else {}
        resp['manual_result'] = manual_info.get('result')
    return jsonify(resp)


@app.route('/generate_report')
def generate_report():
    bound_ip = session.get('bound_ip')
    if not bound_ip:
        return jsonify({'success': False, 'error': '未绑定IP'})
    
    try:
        from script.pdf import RobotTestReport
        import tempfile
        
        data = read_result_file(bound_ip)
        if not data:
            return jsonify({'success': False, 'error': '未找到测试数据'})
        if not isinstance(data, dict):
            data = {}
        data['report_filters'] = build_report_filters(data)
        
        # 使用临时文件生成PDF
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_file:
            temp_filename = temp_file.name
        
        report = RobotTestReport(data)
        report.generate(temp_filename)
        
        # 发送文件并在发送后删除
        response = send_file(
            temp_filename,
            as_attachment=True,
            download_name=f"机器人测试报告_{bound_ip}.pdf",
            mimetype='application/pdf'
        )
        
        # 注册回调函数，在响应完成后删除临时文件
        @response.call_on_close
        def cleanup():
            import os
            if os.path.exists(temp_filename):
                try:
                    os.unlink(temp_filename)
                except:
                    pass
        
        return response
    except Exception as e:
        log.error(f"生成报告出错: {str(e)}")
        import traceback
        log.error(f"堆栈: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': f"{str(e)}"})


@app.route('/image/<image_name>')
def serve_image(image_name):
    try:
        return send_from_directory(os.path.join(app.root_path, 'image'), image_name)
    except Exception as e:
        log.error(f"提供图片失败: {str(e)}")
        return jsonify({'success': False, 'error': '图片不存在'}), 404

@app.route('/get_ptz_image', methods=['GET'])
def get_ptz_image():
    bound_ip = session.get('bound_ip')
    if not bound_ip:
        return jsonify({'success': False, 'error': '未绑定IP'})
    
    img_type = request.args.get('type', 'ptz_1')
    
    try:
        # 调用下载图片的函数
        image_paths = download_latest_image(bound_ip)
        
        if not image_paths:
            return jsonify({'success': False, 'error': '未找到符合条件的图片'})
        
        type_mapping = {
            '1': 'ptz_1',
            '2': 'ptz_2',
            'ptz_1': 'ptz_1',
            'ptz_2': 'ptz_2',
            'preset_1': 'preset_1',
            'preset_2': 'preset_2'
        }
        target_type = type_mapping.get(img_type)
        if not target_type:
            return jsonify({'success': False, 'error': '无效的图片类型'})
        image_path = image_paths.get(target_type)
            
        if not image_path:
            return jsonify({'success': False, 'error': f'未找到类型{img_type}的图片'})
        
        # 返回图片文件
        return send_file(image_path, mimetype='image/jpeg')  # 假设图片是JPEG格式，如果是PNG请修改
    except Exception as e:
        log.error(f"获取云台图片出错: {str(e)}")
        import traceback
        log.error(f"堆栈: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': f"{str(e)}"})


@app.route('/get_integrated_file', methods=['GET'])
def get_integrated_file():
    """
    获取集成测试的图片或视频文件
    
    Args:
        type: 文件类型，可选值：light_photo, light_video, thermal_photo, thermal_video
    
    Returns:
        文件内容
    """
    bound_ip = session.get('bound_ip')
    if not bound_ip:
        return jsonify({'success': False, 'error': '未绑定IP'})
    
    file_type = request.args.get('type', 'light_photo')
    
    # 文件名映射
    filename_map = {
        'light_photo': '可见光拍照.jpg',
        'light_video': '可见光录像.mp4',
        'thermal_photo': '热成像拍照.jpg',
        'thermal_video': '热成像录像.mp4'
    }
    
    filename = filename_map.get(file_type)
    if not filename:
        return jsonify({'success': False, 'error': '无效的文件类型'})
    
    try:
        # 获取 SN 号
        def get_sn():
            try:
                result_file = os.path.join(BASE_DIR, 'test_record', f"{bound_ip}.json")
                if not os.path.exists(result_file):
                    return "UNKNOWN_SN"
                with open(result_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                sn = data.get("robot_info", {}).get("sn") or "UNKNOWN_SN"
                sn = str(sn).strip()
                for ch in ['\\', '/', ':', '*', '?', '"', '<', '>', '|']:
                    sn = sn.replace(ch, "_")
                return sn if sn else "UNKNOWN_SN"
            except Exception:
                return "UNKNOWN_SN"
        
        sn_name = get_sn()
        filename_with_sn = f"{sn_name}_{filename}"
        
        # 构建文件路径
        file_path = os.path.join(BASE_DIR, 'image_yuntai', filename_with_sn)
        
        if not os.path.exists(file_path):
            # 尝试不带 SN 的旧文件名
            old_file_path = os.path.join(BASE_DIR, 'image_yuntai', filename)
            if os.path.exists(old_file_path):
                file_path = old_file_path
            else:
                return jsonify({'success': False, 'error': f'文件不存在: {filename_with_sn}'})
        
        # 根据文件类型设置处理方式
        if 'video' in file_type:
            # 视频文件处理 - 直接下载而不是播放
            log.info(f"正在返回视频文件供下载: {file_path}")
            return send_file(
                file_path,
                mimetype='video/mp4',
                as_attachment=True,
                download_name=os.path.basename(file_path)
            )
        else:
            # 图片直接返回
            mimetype = 'image/jpeg'
            return send_file(file_path, mimetype=mimetype)
    except Exception as e:
        log.error(f"获取集成测试文件出错: {str(e)}")
        import traceback
        log.error(f"堆栈: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': f"{str(e)}"})


@app.route('/get_temperature_data', methods=['GET'])
def get_temperature_data():
    """
    获取测温任务的温度数据
    
    Returns:
        包含 max_temp 和 min_temp 的 JSON 数据
    """
    bound_ip = session.get('bound_ip')
    if not bound_ip:
        return jsonify({'success': False, 'error': '未绑定IP'})
    
    try:
        # 获取 SN 号
        def get_sn():
            try:
                result_file = os.path.join(BASE_DIR, 'test_record', f"{bound_ip}.json")
                if not os.path.exists(result_file):
                    return "UNKNOWN_SN"
                with open(result_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                sn = data.get("robot_info", {}).get("sn") or "UNKNOWN_SN"
                sn = str(sn).strip()
                for ch in ['\\', '/', ':', '*', '?', '"', '<', '>', '|']:
                    sn = sn.replace(ch, "_")
                return sn if sn else "UNKNOWN_SN"
            except Exception:
                return "UNKNOWN_SN"
        
        sn_name = get_sn()
        filename_with_sn = f"{sn_name}_cewen.json"
        
        # 构建文件路径
        file_path = os.path.join(BASE_DIR, 'image_yuntai', filename_with_sn)
        
        if not os.path.exists(file_path):
            # 尝试不带 SN 的旧文件名
            old_file_path = os.path.join(BASE_DIR, 'image_yuntai', 'cewen.json')
            if os.path.exists(old_file_path):
                file_path = old_file_path
                filename_with_sn = 'cewen.json'
            else:
                return jsonify({'success': False, 'error': '暂无温度数据，请先执行测温任务'})
        
        # 读取温度数据
        with open(file_path, 'r', encoding='utf-8') as f:
            temp_data = json.load(f)
        
        return jsonify({
            'success': True,
            'data': temp_data,
            'file_name': filename_with_sn
        })
    except Exception as e:
        log.error(f"获取温度数据出错: {str(e)}")
        import traceback
        log.error(f"堆栈: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': f"{str(e)}"})


@app.route('/get_temperature', methods=['GET'])
def get_temperature_api():
    bound_ip = session.get('bound_ip')
    if not bound_ip:
        return jsonify({'success': False, 'error': '未绑定IP'})
    
    try:
        from script.compass_request import get_temperature
        temperature_data = get_temperature(bound_ip)
        
        if temperature_data:
            return jsonify({'success': True, 'temperature': temperature_data})
        else:
            return jsonify({'success': False, 'error': '未获取到温度数据'})
    except Exception as e:
        log.error(f"获取温度数据出错: {str(e)}")
        import traceback
        log.error(f"堆栈: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': f"{str(e)}"})


@app.route('/admin')
def admin():
    return render_template('admin.html')

@app.route('/admin/get_testcases', methods=['GET'])
def get_testcases():
    model_filter = request.args.get('model', '')
    testcases = []
    
    models = ['MS', 'MR', 'TW', 'X310', 'X320', 'HSR']
    
    for model in models:
        if model_filter and model != model_filter:
            continue
        
        config_path = os.path.join('config', model, 'manual_tests.json')
        if os.path.exists(config_path):
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    for tc in data.get('manual_tests', []):
                        testcases.append({
                            'model': model,
                            'name': tc.get('name', ''),
                            'description': tc.get('description', ''),
                            'expected_result': tc.get('expected_result', ''),
                            'confirm_content': tc.get('confirm_content', ''),
                            'image': tc.get('image', '')
                        })
            except Exception as e:
                log.error(f"读取{config_path}失败: {str(e)}")
    
    return jsonify({'success': True, 'testcases': testcases})

@app.route('/admin/upload_testcase', methods=['POST'])
def upload_testcase():
    try:
        model = request.form.get('model')
        name = request.form.get('name')
        description = request.form.get('description')
        expected_result = request.form.get('expected_result', '')
        confirm_content = request.form.get('confirm_content', '')
        image = request.files.get('image')
        
        if not model or not name or not description:
            return jsonify({'success': False, 'error': '缺少必要参数'})
        
        config_path = os.path.join('config', model, 'manual_tests.json')
        
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        else:
            data = {'manual_tests': []}
        
        new_testcase = {
            'name': name,
            'description': description.split('\n') if '\n' in description else description,
            'expected_result': expected_result,
            'confirm_content': confirm_content
        }
        
        if image:
            ext = os.path.splitext(image.filename)[1]
            unique_name = str(uuid.uuid4()) + ext
            image_path = os.path.join('static', 'images', unique_name)
            image.save(image_path)
            new_testcase['image'] = unique_name
        
        data['manual_tests'].append(new_testcase)
        
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        return jsonify({'success': True})
    
    except Exception as e:
        log.error(f"上传测试用例失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/update_testcase', methods=['POST'])
def update_testcase():
    try:
        model = request.form.get('model')
        index = int(request.form.get('index'))
        name = request.form.get('name')
        description = request.form.get('description')
        expected_result = request.form.get('expected_result', '')
        confirm_content = request.form.get('confirm_content', '')
        image = request.files.get('image')
        
        if not model or name is None or description is None:
            return jsonify({'success': False, 'error': '缺少必要参数'})
        
        config_path = os.path.join('config', model, 'manual_tests.json')
        
        if not os.path.exists(config_path):
            return jsonify({'success': False, 'error': '配置文件不存在'})
        
        with open(config_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        if index < 0 or index >= len(data.get('manual_tests', [])):
            return jsonify({'success': False, 'error': '索引超出范围'})
        
        data['manual_tests'][index]['name'] = name
        data['manual_tests'][index]['description'] = description.split('\n') if '\n' in description else description
        data['manual_tests'][index]['expected_result'] = expected_result
        data['manual_tests'][index]['confirm_content'] = confirm_content
        
        if image:
            ext = os.path.splitext(image.filename)[1]
            unique_name = str(uuid.uuid4()) + ext
            image_path = os.path.join('static', 'images', unique_name)
            image.save(image_path)
            data['manual_tests'][index]['image'] = unique_name
        
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        return jsonify({'success': True})
    
    except Exception as e:
        log.error(f"更新测试用例失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/move_testcase', methods=['POST'])
def move_testcase():
    try:
        data = request.get_json()
        model = data.get('model')
        index = data.get('index')
        direction = data.get('direction')
        
        if not model or index is None or not direction:
            return jsonify({'success': False, 'error': '缺少必要参数'})
        
        config_path = os.path.join('config', model, 'manual_tests.json')
        
        if not os.path.exists(config_path):
            return jsonify({'success': False, 'error': '配置文件不存在'})
        
        with open(config_path, 'r', encoding='utf-8') as f:
            config_data = json.load(f)
        
        manual_tests = config_data.get('manual_tests', [])
        
        if index < 0 or index >= len(manual_tests):
            return jsonify({'success': False, 'error': '索引超出范围'})
        
        if direction == 'up' and index > 0:
            manual_tests[index], manual_tests[index - 1] = manual_tests[index - 1], manual_tests[index]
        elif direction == 'down' and index < len(manual_tests) - 1:
            manual_tests[index], manual_tests[index + 1] = manual_tests[index + 1], manual_tests[index]
        else:
            return jsonify({'success': False, 'error': '无法移动到该位置'})
        
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(config_data, f, ensure_ascii=False, indent=2)
        
        return jsonify({'success': True})
    
    except Exception as e:
        log.error(f"移动测试用例失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/get_version_configs', methods=['GET'])
def get_version_configs():
    try:
        configs = []
        models = ['MS', 'MR', 'TW', 'X310', 'X320', 'HSR']
        
        for model in models:
            config_path = os.path.join('config', model, 'version_config.json')
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    configs.append({
                        'model': model,
                        'pilot_version': data.get('pilot_version', ''),
                        'compass_version': data.get('compass_version', ''),
                        'rcc_base_version': data.get('rcc_base_version', ''),
                        'image_version': data.get('image_version', '')
                    })
        
        return jsonify({'success': True, 'configs': configs})
    
    except Exception as e:
        log.error(f"获取版本配置失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/get_version_config', methods=['GET'])
def get_version_config():
    try:
        model = request.args.get('model')
        if not model:
            return jsonify({'success': False, 'error': '缺少机型参数'})
        
        config_path = os.path.join('config', model, 'version_config.json')
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return jsonify({'success': True, 'config': data})
        else:
            return jsonify({'success': True, 'config': {}})
    
    except Exception as e:
        log.error(f"获取版本配置失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/save_version_config', methods=['POST'])
def save_version_config():
    try:
        data = request.get_json()
        model = data.get('model')
        pilot_version = data.get('pilot_version', '')
        compass_version = data.get('compass_version', '')
        rcc_base_version = data.get('rcc_base_version', '')
        image_version = data.get('image_version', '')
        
        if not model:
            return jsonify({'success': False, 'error': '缺少机型参数'})
        
        config_dir = os.path.join('config', model)
        if not os.path.exists(config_dir):
            os.makedirs(config_dir)
        
        config_path = os.path.join(config_dir, 'version_config.json')
        
        config_data = {
            'pilot_version': pilot_version,
            'compass_version': compass_version,
            'rcc_base_version': rcc_base_version,
            'image_version': image_version
        }
        
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(config_data, f, ensure_ascii=False, indent=2)
        
        return jsonify({'success': True})
    
    except Exception as e:
        log.error(f"保存版本配置失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/delete_version_config', methods=['POST'])
def delete_version_config():
    try:
        data = request.get_json()
        model = data.get('model')
        
        if not model:
            return jsonify({'success': False, 'error': '缺少机型参数'})
        
        config_path = os.path.join('config', model, 'version_config.json')
        if os.path.exists(config_path):
            os.remove(config_path)
        
        return jsonify({'success': True})
    
    except Exception as e:
        log.error(f"删除版本配置失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/delete_testcase', methods=['POST'])
def delete_testcase():
    try:
        data = request.get_json()
        model = data.get('model')
        index = data.get('index')
        
        if not model or index is None:
            return jsonify({'success': False, 'error': '缺少必要参数'})
        
        config_path = os.path.join('config', model, 'manual_tests.json')
        
        if not os.path.exists(config_path):
            return jsonify({'success': False, 'error': '配置文件不存在'})
        
        with open(config_path, 'r', encoding='utf-8') as f:
            config_data = json.load(f)
        
        if index < 0 or index >= len(config_data.get('manual_tests', [])):
            return jsonify({'success': False, 'error': '索引超出范围'})
        
        config_data['manual_tests'].pop(index)
        
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(config_data, f, ensure_ascii=False, indent=2)
        
        return jsonify({'success': True})
    
    except Exception as e:
        log.error(f"删除测试用例失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/batch_delete_testcases', methods=['POST'])
def batch_delete_testcases():
    try:
        data = request.get_json()
        items = data.get('items', [])
        
        if not items or len(items) == 0:
            return jsonify({'success': False, 'error': '请选择要删除的测试用例'})
        
        deleted = 0
        
        for item in items:
            model = item.get('model')
            index = item.get('index')
            
            if not model or index is None:
                continue
            
            config_path = os.path.join('config', model, 'manual_tests.json')
            
            if not os.path.exists(config_path):
                continue
            
            with open(config_path, 'r', encoding='utf-8') as f:
                config_data = json.load(f)
            
            if index >= 0 and index < len(config_data.get('manual_tests', [])):
                config_data['manual_tests'].pop(index)
                
                with open(config_path, 'w', encoding='utf-8') as f:
                    json.dump(config_data, f, ensure_ascii=False, indent=2)
                
                deleted += 1
        
        return jsonify({'success': True, 'deleted': deleted})
    
    except Exception as e:
        log.error(f"批量删除测试用例失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/clear_all_testcases', methods=['POST'])
def clear_all_testcases():
    try:
        data = request.get_json()
        model = data.get('model')
        
        if not model:
            return jsonify({'success': False, 'error': '缺少机型参数'})
        
        config_path = os.path.join('config', model, 'manual_tests.json')
        
        if not os.path.exists(config_path):
            return jsonify({'success': False, 'error': '配置文件不存在'})
        
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump({'manual_tests': []}, f, ensure_ascii=False, indent=2)
        
        return jsonify({'success': True})
    
    except Exception as e:
        log.error(f"清空测试用例失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/upload_excel', methods=['POST'])
def upload_excel():
    temp_path = None
    wb = None
    try:
        model = request.form.get('model')
        excel_file = request.files.get('excel_file')
        
        if not model:
            return jsonify({'success': False, 'error': '请选择机型'})
        if not excel_file:
            return jsonify({'success': False, 'error': '请选择Excel文件'})
        
        filename = excel_file.filename
        if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
            return jsonify({'success': False, 'error': '请上传Excel文件（.xlsx或.xls格式）'})
        
        temp_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp')
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)
        
        temp_path = os.path.join(temp_dir, str(uuid.uuid4()) + '_' + filename)
        excel_file.save(temp_path)
        
        wb = load_workbook(temp_path, read_only=True, data_only=True)
        
        name_col = -1
        desc_col = -1
        expected_col = -1
        header_row = -1
        ws = None
        
        for sheet_name in wb.sheetnames:
            current_ws = wb[sheet_name]
            current_name_col = -1
            current_desc_col = -1
            current_expected_col = -1
            current_header_row = -1
            
            for row in range(1, min(20, current_ws.max_row + 1)):
                for col in range(1, current_ws.max_column + 1):
                    cell_value = current_ws.cell(row=row, column=col).value
                    if cell_value:
                        cell_str = str(cell_value).strip()
                        if current_name_col == -1 and ('测试用例名称' in cell_str or '测试用例' in cell_str or 'name' in cell_str.lower()):
                            current_name_col = col
                            current_header_row = row
                        elif current_desc_col == -1 and ('测试用例描述' in cell_str or '描述' in cell_str or 'description' in cell_str.lower() or '说明' in cell_str or '步骤' in cell_str or '内容' in cell_str):
                            current_desc_col = col
                        elif current_expected_col == -1 and ('预期结果' in cell_str or '期望结果' in cell_str or 'expected' in cell_str.lower() or 'result' in cell_str.lower() or '结果' in cell_str):
                            current_expected_col = col
            
            if current_name_col != -1 and current_desc_col != -1:
                name_col = current_name_col
                desc_col = current_desc_col
                expected_col = current_expected_col
                header_row = current_header_row
                ws = current_ws
                break
        
        if name_col == -1:
            return jsonify({'success': False, 'error': '未找到"测试用例名称"列，请确保Excel包含"测试用例名称"或"名称"列'})
        if desc_col == -1:
            return jsonify({'success': False, 'error': '未找到"测试用例描述"列，请确保Excel包含"测试用例描述"或"描述"列'})
        
        config_path = os.path.join('config', model, 'manual_tests.json')
        
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        else:
            data = {'manual_tests': []}
        
        count = 0
        skipped = 0
        
        for row in range(header_row + 1, ws.max_row + 1):
            name = ws.cell(row=row, column=name_col).value
            description = ws.cell(row=row, column=desc_col).value
            
            if name:
                name_str = str(name).strip()
            else:
                name_str = ''
            
            if description:
                desc_str = str(description).strip()
            else:
                desc_str = ''
            
            if expected_col != -1:
                expected = ws.cell(row=row, column=expected_col).value
                if expected and str(expected).strip():
                    if desc_str:
                        desc_str = desc_str + ' ' + str(expected).strip()
                    else:
                        desc_str = str(expected).strip()
            
            if not name_str:
                skipped += 1
                continue
            
            if name_str.startswith('='):
                skipped += 1
                continue
            
            if 'Category' in name_str or 'Feature Test' in name_str or '类别介绍' in name_str:
                skipped += 1
                continue
            
            if len(name_str) < 2:
                skipped += 1
                continue
            
            data['manual_tests'].append({
                'name': name_str,
                'description': desc_str
            })
            count += 1
        
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        return jsonify({'success': True, 'count': count, 'skipped': skipped})
    
    except Exception as e:
        log.error(f"批量导入Excel失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})
    
    finally:
        if wb:
            try:
                wb.close()
            except:
                pass
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except:
                pass

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
