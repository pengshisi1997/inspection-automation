import json
import os
import sys
import subprocess
import uuid

from flask import Blueprint, jsonify, render_template, request

from config.model_config import MODEL_OPTIONS
from log import Logger

try:
    from openpyxl import load_workbook
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
log = Logger()
admin_bp = Blueprint('admin', __name__)

@admin_bp.route('/admin')
def admin():
    return render_template('admin.html')

@admin_bp.route('/admin/get_testcases', methods=['GET'])
def get_testcases():
    model_filter = request.args.get('model', '')
    testcases = []
    
    models = MODEL_OPTIONS
    
    for model in models:
        if model_filter and model != model_filter:
            continue
        
        config_path = os.path.join('config', model, 'manual_tests.json')
        if os.path.exists(config_path):
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    for tc in data.get('manual_tests', []):
                        testcases.append({
                            'model': model,
                            'name': tc.get('name', ''),
                            'description': tc.get('description', ''),
                            'expected_result': tc.get('expected_result', ''),
                            'confirm_content': tc.get('confirm_content', ''),
                            'image': tc.get('image', '')
                        })
            except Exception as e:
                log.error(f"读取{config_path}失败: {str(e)}")
    
    return jsonify({'success': True, 'testcases': testcases})

@admin_bp.route('/admin/upload_testcase', methods=['POST'])
def upload_testcase():
    try:
        model = request.form.get('model')
        name = request.form.get('name')
        description = request.form.get('description')
        expected_result = request.form.get('expected_result', '')
        confirm_content = request.form.get('confirm_content', '')
        image = request.files.get('image')
        
        if not model or not name or not description:
            return jsonify({'success': False, 'error': '缺少必要参数'})
        
        config_path = os.path.join('config', model, 'manual_tests.json')
        
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        else:
            data = {'manual_tests': []}
        
        new_testcase = {
            'name': name,
            'description': description.split('\n') if '\n' in description else description,
            'expected_result': expected_result,
            'confirm_content': confirm_content
        }
        
        if image:
            ext = os.path.splitext(image.filename)[1]
            unique_name = str(uuid.uuid4()) + ext
            image_path = os.path.join('static', 'images', unique_name)
            image.save(image_path)
            new_testcase['image'] = unique_name
        
        data['manual_tests'].append(new_testcase)
        
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        return jsonify({'success': True})
    
    except Exception as e:
        log.error(f"上传测试用例失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@admin_bp.route('/admin/update_testcase', methods=['POST'])
def update_testcase():
    try:
        model = request.form.get('model')
        index = int(request.form.get('index'))
        name = request.form.get('name')
        description = request.form.get('description')
        expected_result = request.form.get('expected_result', '')
        confirm_content = request.form.get('confirm_content', '')
        image = request.files.get('image')
        
        if not model or name is None or description is None:
            return jsonify({'success': False, 'error': '缺少必要参数'})
        
        config_path = os.path.join('config', model, 'manual_tests.json')
        
        if not os.path.exists(config_path):
            return jsonify({'success': False, 'error': '配置文件不存在'})
        
        with open(config_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        if index < 0 or index >= len(data.get('manual_tests', [])):
            return jsonify({'success': False, 'error': '索引超出范围'})
        
        data['manual_tests'][index]['name'] = name
        data['manual_tests'][index]['description'] = description.split('\n') if '\n' in description else description
        data['manual_tests'][index]['expected_result'] = expected_result
        data['manual_tests'][index]['confirm_content'] = confirm_content
        
        if image:
            ext = os.path.splitext(image.filename)[1]
            unique_name = str(uuid.uuid4()) + ext
            image_path = os.path.join('static', 'images', unique_name)
            image.save(image_path)
            data['manual_tests'][index]['image'] = unique_name
        
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        return jsonify({'success': True})
    
    except Exception as e:
        log.error(f"更新测试用例失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@admin_bp.route('/admin/move_testcase', methods=['POST'])
def move_testcase():
    try:
        data = request.get_json()
        model = data.get('model')
        index = data.get('index')
        direction = data.get('direction')
        
        if not model or index is None or not direction:
            return jsonify({'success': False, 'error': '缺少必要参数'})
        
        config_path = os.path.join('config', model, 'manual_tests.json')
        
        if not os.path.exists(config_path):
            return jsonify({'success': False, 'error': '配置文件不存在'})
        
        with open(config_path, 'r', encoding='utf-8') as f:
            config_data = json.load(f)
        
        manual_tests = config_data.get('manual_tests', [])
        
        if index < 0 or index >= len(manual_tests):
            return jsonify({'success': False, 'error': '索引超出范围'})
        
        if direction == 'up' and index > 0:
            manual_tests[index], manual_tests[index - 1] = manual_tests[index - 1], manual_tests[index]
        elif direction == 'down' and index < len(manual_tests) - 1:
            manual_tests[index], manual_tests[index + 1] = manual_tests[index + 1], manual_tests[index]
        else:
            return jsonify({'success': False, 'error': '无法移动到该位置'})
        
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(config_data, f, ensure_ascii=False, indent=2)
        
        return jsonify({'success': True})
    
    except Exception as e:
        log.error(f"移动测试用例失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@admin_bp.route('/admin/get_version_configs', methods=['GET'])
def get_version_configs():
    try:
        configs = []
        models = MODEL_OPTIONS
        
        for model in models:
            config_path = os.path.join('config', model, 'version_config.json')
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    configs.append({
                        'model': model,
                        'pilot_version': data.get('pilot_version', ''),
                        'compass_version': data.get('compass_version', ''),
                        'rcc_base_version': data.get('rcc_base_version', ''),
                        'image_version': data.get('image_version', '')
                    })
        
        return jsonify({'success': True, 'configs': configs})
    
    except Exception as e:
        log.error(f"获取版本配置失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@admin_bp.route('/admin/get_version_config', methods=['GET'])
def get_version_config():
    try:
        model = request.args.get('model')
        if not model:
            return jsonify({'success': False, 'error': '缺少机型参数'})
        
        config_path = os.path.join('config', model, 'version_config.json')
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return jsonify({'success': True, 'config': data})
        else:
            return jsonify({'success': True, 'config': {}})
    
    except Exception as e:
        log.error(f"获取版本配置失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@admin_bp.route('/admin/save_version_config', methods=['POST'])
def save_version_config():
    try:
        data = request.get_json()
        model = data.get('model')
        pilot_version = data.get('pilot_version', '')
        compass_version = data.get('compass_version', '')
        rcc_base_version = data.get('rcc_base_version', '')
        image_version = data.get('image_version', '')
        
        if not model:
            return jsonify({'success': False, 'error': '缺少机型参数'})
        
        config_dir = os.path.join('config', model)
        if not os.path.exists(config_dir):
            os.makedirs(config_dir)
        
        config_path = os.path.join(config_dir, 'version_config.json')
        
        config_data = {
            'pilot_version': pilot_version,
            'compass_version': compass_version,
            'rcc_base_version': rcc_base_version,
            'image_version': image_version
        }
        
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(config_data, f, ensure_ascii=False, indent=2)
        
        return jsonify({'success': True})
    
    except Exception as e:
        log.error(f"保存版本配置失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@admin_bp.route('/admin/delete_version_config', methods=['POST'])
def delete_version_config():
    try:
        data = request.get_json()
        model = data.get('model')
        
        if not model:
            return jsonify({'success': False, 'error': '缺少机型参数'})
        
        config_path = os.path.join('config', model, 'version_config.json')
        if os.path.exists(config_path):
            os.remove(config_path)
        
        return jsonify({'success': True})
    
    except Exception as e:
        log.error(f"删除版本配置失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@admin_bp.route('/admin/delete_testcase', methods=['POST'])
def delete_testcase():
    try:
        data = request.get_json()
        model = data.get('model')
        index = data.get('index')
        
        if not model or index is None:
            return jsonify({'success': False, 'error': '缺少必要参数'})
        
        config_path = os.path.join('config', model, 'manual_tests.json')
        
        if not os.path.exists(config_path):
            return jsonify({'success': False, 'error': '配置文件不存在'})
        
        with open(config_path, 'r', encoding='utf-8') as f:
            config_data = json.load(f)
        
        if index < 0 or index >= len(config_data.get('manual_tests', [])):
            return jsonify({'success': False, 'error': '索引超出范围'})
        
        config_data['manual_tests'].pop(index)
        
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(config_data, f, ensure_ascii=False, indent=2)
        
        return jsonify({'success': True})
    
    except Exception as e:
        log.error(f"删除测试用例失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@admin_bp.route('/admin/batch_delete_testcases', methods=['POST'])
def batch_delete_testcases():
    try:
        data = request.get_json()
        items = data.get('items', [])
        
        if not items or len(items) == 0:
            return jsonify({'success': False, 'error': '请选择要删除的测试用例'})
        
        deleted = 0
        
        for item in items:
            model = item.get('model')
            index = item.get('index')
            
            if not model or index is None:
                continue
            
            config_path = os.path.join('config', model, 'manual_tests.json')
            
            if not os.path.exists(config_path):
                continue
            
            with open(config_path, 'r', encoding='utf-8') as f:
                config_data = json.load(f)
            
            if index >= 0 and index < len(config_data.get('manual_tests', [])):
                config_data['manual_tests'].pop(index)
                
                with open(config_path, 'w', encoding='utf-8') as f:
                    json.dump(config_data, f, ensure_ascii=False, indent=2)
                
                deleted += 1
        
        return jsonify({'success': True, 'deleted': deleted})
    
    except Exception as e:
        log.error(f"批量删除测试用例失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@admin_bp.route('/admin/clear_all_testcases', methods=['POST'])
def clear_all_testcases():
    try:
        data = request.get_json()
        model = data.get('model')
        
        if not model:
            return jsonify({'success': False, 'error': '缺少机型参数'})
        
        config_path = os.path.join('config', model, 'manual_tests.json')
        
        if not os.path.exists(config_path):
            return jsonify({'success': False, 'error': '配置文件不存在'})
        
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump({'manual_tests': []}, f, ensure_ascii=False, indent=2)
        
        return jsonify({'success': True})
    
    except Exception as e:
        log.error(f"清空测试用例失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@admin_bp.route('/admin/upload_excel', methods=['POST'])
def upload_excel():
    temp_path = None
    wb = None
    try:
        model = request.form.get('model')
        excel_file = request.files.get('excel_file')
        
        if not model:
            return jsonify({'success': False, 'error': '请选择机型'})
        if not excel_file:
            return jsonify({'success': False, 'error': '请选择Excel文件'})
        
        filename = excel_file.filename
        if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
            return jsonify({'success': False, 'error': '请上传Excel文件（.xlsx或.xls格式）'})
        
        temp_dir = os.path.join(BASE_DIR, 'temp')
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)
        
        temp_path = os.path.join(temp_dir, str(uuid.uuid4()) + '_' + filename)
        excel_file.save(temp_path)
        
        wb = load_workbook(temp_path, read_only=True, data_only=True)
        
        name_col = -1
        desc_col = -1
        expected_col = -1
        header_row = -1
        ws = None
        
        for sheet_name in wb.sheetnames:
            current_ws = wb[sheet_name]
            current_name_col = -1
            current_desc_col = -1
            current_expected_col = -1
            current_header_row = -1
            
            for row in range(1, min(20, current_ws.max_row + 1)):
                for col in range(1, current_ws.max_column + 1):
                    cell_value = current_ws.cell(row=row, column=col).value
                    if cell_value:
                        cell_str = str(cell_value).strip()
                        if current_name_col == -1 and ('测试用例名称' in cell_str or '测试用例' in cell_str or 'name' in cell_str.lower()):
                            current_name_col = col
                            current_header_row = row
                        elif current_desc_col == -1 and ('测试用例描述' in cell_str or '描述' in cell_str or 'description' in cell_str.lower() or '说明' in cell_str or '步骤' in cell_str or '内容' in cell_str):
                            current_desc_col = col
                        elif current_expected_col == -1 and ('预期结果' in cell_str or '期望结果' in cell_str or 'expected' in cell_str.lower() or 'result' in cell_str.lower() or '结果' in cell_str):
                            current_expected_col = col
            
            if current_name_col != -1 and current_desc_col != -1:
                name_col = current_name_col
                desc_col = current_desc_col
                expected_col = current_expected_col
                header_row = current_header_row
                ws = current_ws
                break
        
        if name_col == -1:
            return jsonify({'success': False, 'error': '未找到"测试用例名称"列，请确保Excel包含"测试用例名称"或"名称"列'})
        if desc_col == -1:
            return jsonify({'success': False, 'error': '未找到"测试用例描述"列，请确保Excel包含"测试用例描述"或"描述"列'})
        
        config_path = os.path.join('config', model, 'manual_tests.json')
        
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        else:
            data = {'manual_tests': []}
        
        count = 0
        skipped = 0
        
        for row in range(header_row + 1, ws.max_row + 1):
            name = ws.cell(row=row, column=name_col).value
            description = ws.cell(row=row, column=desc_col).value
            
            if name:
                name_str = str(name).strip()
            else:
                name_str = ''
            
            if description:
                desc_str = str(description).strip()
            else:
                desc_str = ''
            
            if expected_col != -1:
                expected = ws.cell(row=row, column=expected_col).value
                if expected and str(expected).strip():
                    if desc_str:
                        desc_str = desc_str + ' ' + str(expected).strip()
                    else:
                        desc_str = str(expected).strip()
            
            if not name_str:
                skipped += 1
                continue
            
            if name_str.startswith('='):
                skipped += 1
                continue
            
            if 'Category' in name_str or 'Feature Test' in name_str or '类别介绍' in name_str:
                skipped += 1
                continue
            
            if len(name_str) < 2:
                skipped += 1
                continue
            
            data['manual_tests'].append({
                'name': name_str,
                'description': desc_str
            })
            count += 1
        
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        return jsonify({'success': True, 'count': count, 'skipped': skipped})
    
    except Exception as e:
        log.error(f"批量导入Excel失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})
    
    finally:
        if wb:
            try:
                wb.close()
            except:
                pass
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except:
                pass

